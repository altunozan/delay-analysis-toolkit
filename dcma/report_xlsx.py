"""Excel report builder for DCMA assessment results.

Produces a formatted workbook mirroring the UI scorecard:
  - "Summary" sheet: project header + color-coded 14-check scorecard table
    with the DCMA rationale for each target.
  - One detail sheet per check that has affected activities.
  - Optional "AI Narrative" sheet when a generated narrative is supplied.

UI-independent: returns bytes, so it can serve Streamlit downloads or CLI use.
"""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .checks import CheckResult, CheckStatus
from .rationale import CHECK_RATIONALE
from .xer_parser import XerData

# Scorecard colors (match the UI palette).
STATUS_FILLS = {
    CheckStatus.PASS: PatternFill("solid", fgColor="C6EFCE"),
    CheckStatus.FAIL: PatternFill("solid", fgColor="FFC7CE"),
    CheckStatus.NA: PatternFill("solid", fgColor="D9D9D9"),
}
STATUS_FONTS = {
    CheckStatus.PASS: Font(color="1A7F37", bold=True),
    CheckStatus.FAIL: Font(color="9C0006", bold=True),
    CheckStatus.NA: Font(color="6E7781", bold=True),
}

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=16, bold=True, color="1F3864")
THIN_BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)
WRAP = Alignment(wrap_text=True, vertical="top")


def _autofit(ws, widths: dict[int, int]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _sheet_name(number: int, name: str) -> str:
    """Excel sheet names are capped at 31 chars and bar some characters."""
    clean = name.replace("/", "-").replace("\\", "-")
    return f"{number:02d} {clean}"[:31]


def build_xlsx_report(
    data: XerData,
    results: list[CheckResult],
    narrative: str | None = None,
    trace=None,
) -> bytes:
    """``trace`` (dcma.trace.DCMATrace, optional) adds traceback sheets."""
    wb = Workbook()

    # ------------------------------------------------------------------ #
    # Summary sheet
    # ------------------------------------------------------------------ #
    ws = wb.active
    ws.title = "Summary"

    proj = data.project
    ws["A1"] = "DCMA 14-Point Schedule Assessment Report"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")

    meta = [
        ("Project", proj.short_name if proj else "—"),
        ("Data Date", proj.data_date.strftime("%Y-%m-%d") if proj and proj.data_date else "—"),
        ("Activities", len(data.tasks)),
        ("Relationships", len(data.relationships)),
        ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    row = 3
    for label, value in meta:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    passed = sum(1 for r in results if r.status == CheckStatus.PASS)
    failed = sum(1 for r in results if r.status == CheckStatus.FAIL)
    na = sum(1 for r in results if r.status == CheckStatus.NA)
    scored = passed + failed
    score_pct = (passed / scored * 100.0) if scored else 0.0
    ws.cell(row=row, column=1, value="Overall Score").font = Font(bold=True)
    ws.cell(row=row, column=2,
            value=f"{passed}/{scored} scored checks passed ({score_pct:.0f}%) — "
                  f"{failed} failed, {na} N/A")
    row += 2

    # Scorecard table
    headers = ["#", "Check", "Status", "Metric", "Value", "Threshold",
               "Affected", "Why This Target Matters (DCMA)"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = THIN_BORDER
    row += 1

    for r in results:
        values = [
            r.number, r.name, r.status.value, r.metric_label, r.metric_value,
            r.threshold, r.affected_count, CHECK_RATIONALE.get(r.number, ""),
        ]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = THIN_BORDER
            c.alignment = WRAP
            if col == 3:  # Status column gets the color treatment
                c.fill = STATUS_FILLS[r.status]
                c.font = STATUS_FONTS[r.status]
                c.alignment = Alignment(horizontal="center", vertical="top")
        row += 1

    _autofit(ws, {1: 4, 2: 20, 3: 9, 4: 38, 5: 26, 6: 14, 7: 9, 8: 60})
    ws.freeze_panes = f"A{row - len(results)}"

    # ------------------------------------------------------------------ #
    # Detail sheets — one per check with affected activities
    # ------------------------------------------------------------------ #
    for r in results:
        if not r.detail_rows:
            continue
        dws = wb.create_sheet(_sheet_name(r.number, r.name))
        dws["A1"] = f"Check {r.number}: {r.name} — {r.status.value}"
        dws["A1"].font = Font(size=13, bold=True, color="1F3864")
        dws["A2"] = r.summary
        dws["A2"].alignment = WRAP
        dws["A3"] = f"Target: {r.threshold}   ·   Result: {r.metric_value}"
        dws["A3"].font = Font(italic=True)

        cols = list(r.detail_rows[0].keys())
        hrow = 5
        for col, h in enumerate(cols, start=1):
            c = dws.cell(row=hrow, column=col, value=h)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.border = THIN_BORDER
        for i, detail in enumerate(r.detail_rows, start=hrow + 1):
            for col, key in enumerate(cols, start=1):
                c = dws.cell(row=i, column=col, value=detail.get(key))
                c.border = THIN_BORDER
        _autofit(dws, {i: 28 for i in range(1, len(cols) + 1)})
        dws.freeze_panes = f"A{hrow + 1}"

    # ------------------------------------------------------------------ #
    # Traceback sheets (optional)
    # ------------------------------------------------------------------ #
    if trace is not None and trace.chain and trace.chain.steps:
        c = trace.chain
        tws = wb.create_sheet("Driving Chain")
        tws["A1"] = (f"Driving chain to {c.terminal_code} "
                     f"'{c.terminal_name}' — stored dates & logic")
        tws["A1"].font = Font(size=13, bold=True, color="1F3864")
        tws["A2"] = ("Continuous back to the data date."
                     if c.reaches_data_date else
                     f"BREAKS at {c.break_code}: {c.break_reason}")
        tws["A2"].font = Font(italic=True)
        headers = ["#", "Activity ID", "Activity Name", "Milestone",
                   "Early Start", "Early Finish", "TF (d)",
                   "Driven By (link)", "Constraint(s)"]
        for col, h in enumerate(headers, start=1):
            cell = tws.cell(row=4, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
        for i, s in enumerate(c.steps, start=5):
            vals = [s.seq, s.task_code, s.name,
                    "Yes" if s.is_milestone else "",
                    s.early_start.strftime("%Y-%m-%d") if s.early_start else "",
                    s.early_finish.strftime("%Y-%m-%d") if s.early_finish else "",
                    s.total_float_days, s.link_from_prev, s.constraint]
            for col, v in enumerate(vals, start=1):
                tws.cell(row=i, column=col, value=v).border = THIN_BORDER
        _autofit(tws, {1: 5, 2: 18, 3: 46, 4: 10, 5: 12, 6: 12, 7: 9,
                       8: 24, 9: 30})
        tws.freeze_panes = "A5"

    if trace is not None and trace.float_driver_groups:
        fws = wb.create_sheet("Float Drivers")
        fws["A1"] = ("Negative float traced to its governing constraint "
                     "(Check 5 -> Check 7 causation, stored values)")
        fws["A1"].font = Font(size=13, bold=True, color="1F3864")
        headers = ["Negative-float activities", "Worst TF (d)",
                   "Governing driver", "Kind", "Example trace"]
        for col, h in enumerate(headers, start=1):
            cell = fws.cell(row=3, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
        for i, g in enumerate(trace.float_driver_groups, start=4):
            ex = ""
            if g.example:
                ex = g.example.origin_code
                if g.example.via_codes:
                    ex += " -> " + " -> ".join(g.example.via_codes[:8])
            vals = [g.count, g.worst_tf_days, g.driver_detail,
                    g.driver_kind, ex]
            for col, v in enumerate(vals, start=1):
                cell = fws.cell(row=i, column=col, value=v)
                cell.border = THIN_BORDER
                cell.alignment = WRAP
        _autofit(fws, {1: 12, 2: 12, 3: 56, 4: 18, 5: 52})
        fws.freeze_panes = "A4"

    if trace is not None and trace.offenders:
        ows = wb.create_sheet("Multi-Check Offenders")
        ows["A1"] = ("Activities tripping two or more checks, "
                     "driving-path first")
        ows["A1"].font = Font(size=13, bold=True, color="1F3864")
        headers = ["Activity ID", "Activity Name", "Path position",
                   "Checks tripped"]
        for col, h in enumerate(headers, start=1):
            cell = ows.cell(row=3, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
        for i, o in enumerate(trace.offenders[:500], start=4):
            vals = [o.task_code, o.name, o.band, o.checks_label]
            for col, v in enumerate(vals, start=1):
                ows.cell(row=i, column=col, value=v).border = THIN_BORDER
        _autofit(ows, {1: 18, 2: 52, 3: 14, 4: 16})
        ows.freeze_panes = "A4"

    if trace is not None and (trace.caveats or trace.warnings):
        cws = wb.create_sheet("Traceback Notes")
        cws["A1"] = "Traceback warnings & standing caveats"
        cws["A1"].font = Font(size=13, bold=True, color="1F3864")
        cws.column_dimensions["A"].width = 110
        row2 = 3
        for note in trace.warnings + trace.caveats:
            cell = cws.cell(row=row2, column=1, value="• " + note)
            cell.alignment = WRAP
            row2 += 1

    # ------------------------------------------------------------------ #
    # AI narrative sheet (optional)
    # ------------------------------------------------------------------ #
    if narrative:
        nws = wb.create_sheet("AI Narrative")
        nws["A1"] = "AI-Generated Schedule Analytics Narrative"
        nws["A1"].font = Font(size=13, bold=True, color="1F3864")
        nws.column_dimensions["A"].width = 110
        row = 3
        for para in narrative.split("\n"):
            c = nws.cell(row=row, column=1, value=para)
            c.alignment = WRAP
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
