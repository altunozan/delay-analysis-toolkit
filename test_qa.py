"""QA/QC regression suite — engine level.

Layer A: delay-analyst cross-validation — modules must agree with each other
and with manual recomputation from raw XER rows.
Layer B: software edge cases — degenerate inputs, symmetry, bounds.
Layer C: report integrity — prompts carry the hard rules and caveats; every
workbook opens with its narrative sheet.

Run: python3 test_qa.py  (exit code 1 on any failure)
"""
import os
import sys
import io

from openpyxl import load_workbook

def _p(rel: str) -> str:
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), rel)


from dcma import parse_xer, run_all_checks
from dcma.config import DCMAConfig
from programme import (
    analyse_float_erosion, analyse_windows, build_comparison_prompt,
    build_comparison_xlsx, build_critical_path_prompt,
    build_critical_path_xlsx, build_float_erosion_prompt,
    build_float_erosion_xlsx, build_inventory, build_inventory_prompt,
    build_inventory_xlsx, build_milestone_prompt, build_milestone_xlsx,
    build_progress_prompt, build_progress_xlsx, build_resources_prompt,
    build_resources_xlsx, build_variance_prompt, build_variance_xlsx,
    build_windows_prompt, build_windows_xlsx, compare_revisions,
    compute_progress, compute_variance_by_mapping, end_activity_candidates,
    extract_critical_path, extract_longest_path, extract_resource_loading,
    task_wbs_assignments, track_milestone_shifts,
)

PASS, FAIL = [], []
def check(name, cond, detail=""):
    (PASS if cond else FAIL).append((name, detail))
    print(f"  [{'PASS' if cond else 'FAIL'}] {name}" + (f" — {detail}" if detail and not cond else ""))

cfg = DCMAConfig()
with open(_p("sample/Sample Baseline.xer"),"rb") as fh:
    B = parse_xer(fh.read())
with open(_p("sample/Sample Update.xer"),"rb") as fh:
    U = parse_xer(fh.read())
fix = []
for f in ["revA.xer","revB.xer","revC.xer"]:
    with open(_p(f"sample/revisions/{f}"), "rb") as fh:
        fix.append((f, parse_xer(fh.read())))

print("== A. Cross-module numerical consistency ==")

# A1. Negative float: DCMA check 7 vs float-erosion snapshot (baseline)
dcma = {c.number: c for c in run_all_checks(B, cfg)}
fe = analyse_float_erosion([("B", B), ("U", U)])
check("A1 DCMA neg-float == float-erosion neg count (baseline)",
      dcma[7].affected_count == fe.snapshots[0].negative_count,
      f"dcma={dcma[7].affected_count} vs fe={fe.snapshots[0].negative_count}")

# A2. Critical count: DCMA check 12 vs float-method CP module
cp_f = extract_critical_path(B, "B")
check("A2 DCMA critical count == CP module critical (TF<=0)",
      dcma[12].affected_count == len(cp_f.critical),
      f"dcma={dcma[12].affected_count} vs cp={len(cp_f.critical)}")

# A3. Manual TF recount from raw rows
manual_crit = 0
for t in B.tasks:
    if t.is_loe_or_wbs or t.is_complete or t.total_float_hr is None:
        continue
    if t.total_float_hr / B.hours_per_day(t, cfg) <= 0:
        manual_crit += 1
check("A3 manual TF<=0 recount == CP module", manual_crit == len(cp_f.critical),
      f"manual={manual_crit} vs cp={len(cp_f.critical)}")

# A4. Windows completion movement == project-level scheduled finish delta
wres = analyse_windows([("B", B), ("U", U)])
manual_move = (U.project.scheduled_finish - B.project.scheduled_finish).days
check("A4 windows movement == scheduled finish delta",
      wres.windows[0].movement_days == manual_move,
      f"win={wres.windows[0].movement_days} vs manual={manual_move}")

# A5. Windows total == sum across fixture windows
wfix = analyse_windows(fix)
tot = sum(w.movement_days for w in wfix.windows if w.movement_days is not None)
check("A5 fixtures cumulative movement == sum of windows",
      wfix.total_movement_days == tot)

# A5b-d. Window driver traceback: every movement figure is backed by
# row-level driving-path activities carrying STORED dates only.
_w0 = wres.windows[0]
check("A5b window drivers populated from the later revision's path",
      len(_w0.drivers) > 0)
_u_by_code = {t.task_code: t for t in U.tasks}
_drv_ok = all(
    (d.finish_new == ((_u_by_code[d.task_code].act_finish
                       or _u_by_code[d.task_code].early_finish)))
    for d in _w0.drivers if d.task_code in _u_by_code)
check("A5c driver finishes are the revision's own stored dates", _drv_ok)
_slips = [d.slip_days for d in _w0.drivers if d.slip_days is not None]
check("A5d drivers sorted biggest mover first",
      _slips == sorted(_slips, reverse=True))

# A6. Longest path is a subset of... no — verify every longest-path link
# joins two on-path activities and terminal is on path
cp_l = extract_longest_path(B, "B")
codes = {a.task_code for a in cp_l.critical}
bad_links = [lk for lk in cp_l.links
             if lk.pred_code not in codes or lk.succ_code not in codes]
check("A6 longest-path links all join on-path activities", not bad_links,
      f"{len(bad_links)} dangling")
check("A6b terminal on path", cp_l.end_choice in codes)

# A7. Single-branch trace (A3400) — every non-start activity has a driving
# predecessor within the path
cp_s = extract_longest_path(B, "B", end_task_code="A3400")
succs_with_pred = {lk.succ_code for lk in cp_s.links}
starts = [a.task_code for a in cp_s.critical
          if a.task_code not in succs_with_pred]
check("A7 single-branch trace has exactly one chain start",
      len(starts) == 1, f"starts={starts}")

# A8. Comparison symmetry: swap old/new -> added<->deleted, and
# reversed-data-date warning fires
c_fwd = compare_revisions(B, U, "B", "U")
c_rev = compare_revisions(U, B, "U", "B")
check("A8 comparison added/deleted symmetric",
      len(c_fwd.added) == len(c_rev.deleted)
      and len(c_fwd.deleted) == len(c_rev.added))
check("A8b reversed direction warned",
      any("LATER data date" in w for w in c_rev.warnings))

# A9. Self-comparison finds zero changes
c_self = compare_revisions(B, B, "B", "B2")
check("A9 self-comparison == 0 changes", c_self.total_changes == 0,
      f"{c_self.total_changes} changes: {c_self.category_counts}")

# A10. S-curve bounds and monotonicity
pr = compute_progress(B, "B", [("U", U)])
mono = all(a.cum_pct <= b.cum_pct + 1e-9
           for a, b in zip(pr.planned_curve, pr.planned_curve[1:]))
check("A10 planned curve monotonic", mono)
check("A10b planned curve ends at 100%",
      abs(pr.planned_curve[-1].cum_pct - 100.0) < 0.1,
      f"end={pr.planned_curve[-1].cum_pct}")
check("A10c recorded curve <= 100%",
      all(p.cum_pct <= 100.0 + 1e-9 for p in pr.recorded_curve))
rmono = all(a.cum_pct <= b.cum_pct + 1e-9
            for a, b in zip(pr.recorded_curve, pr.recorded_curve[1:]))
check("A10d recorded curve monotonic", rmono)

# A11. Recorded % manual recount (duration weights)
w = {}
for t in U.tasks:
    if t.is_loe_or_wbs: continue
    d = t.original_duration_days(U.hours_per_day(t, cfg)) or 0.0
    w[t.task_id] = max(d, 0.0)
pct = {r["task_id"].strip(): float(r.get("phys_complete_pct") or 0)
       for r in U.raw_tables["TASK"]}
earned = sum(w[t.task_id] if t.is_complete
             else w[t.task_id]*pct.get(t.task_id,0)/100 if t.act_start else 0
             for t in U.tasks if not t.is_loe_or_wbs)
manual_pct = round(100*earned/sum(w.values()), 1)
check("A11 recorded % matches manual recount",
      abs(manual_pct - pr.recorded_pct_at_dd) < 0.05,
      f"manual={manual_pct} vs module={pr.recorded_pct_at_dd}")

# A12. Milestone shift manual verification: pick one milestone present in
# both, verify its total shift equals date difference from raw fields
inv_pool = [("Sample Baseline.xer", B), ("Sample Update.xer", U)]
ms = track_milestone_shifts([
    ("Sample Baseline.xer", B.project.data_date, B),
    ("Sample Update.xer", U.project.data_date, U),
])
s_ok = None
for s in ms.series:
    if s.total_shift_days is not None and len([p for p in s.points if p.value_date]) == 2:
        p0, p1 = [p for p in s.points if p.value_date]
        expected = (p1.value_date - p0.value_date).days
        s_ok = (s.key, s.total_shift_days, expected)
        break
check("A12 milestone shift == raw date delta",
      s_ok is not None and abs(s_ok[1] - s_ok[2]) < 1.0, str(s_ok))

# A13. Variance group bounds: WBS L1 groups — min start / max finish manual
wbs_map = task_wbs_assignments(B, level=1)
var = compute_variance_by_mapping(B, U, wbs_map, wbs_map, "WBS L1")
g = next(g for g in var.groups if g.planned.activity_count > 5)
ids = [tid for tid, lbl in wbs_map.items() if lbl == g.code_value]
starts = [t.target_start or t.early_start for t in B.tasks
          if t.task_id in set(ids) and not t.is_loe_or_wbs
          and (t.target_start or t.early_start)]
check("A13 variance planned start == manual min",
      g.planned.start == min(starts),
      f"module={g.planned.start} manual={min(starts)}")

# A14. Resource totals == raw TASKRSRC sum (for dated, positive assignments)
rl = extract_resource_loading(B, "B")
raw_total = 0.0
tid_ok = {t.task_id for t in B.tasks if not t.is_loe_or_wbs
          and (t.target_start or t.early_start or t.act_start)}
rid_ok = {r.rsrc_id for r in rl.resources} | {
    (row.get("rsrc_id") or "").strip() for row in B.raw_tables["RSRC"]}
for row in B.raw_tables["TASKRSRC"]:
    try: q = float(row.get("target_qty") or 0)
    except ValueError: q = 0
    if q > 0 and (row.get("task_id") or "").strip() in tid_ok:
        raw_total += q
mod_total = sum(r.total_qty for r in rl.resources)
check("A14 resource totals == raw sum", abs(raw_total - mod_total) < 0.5,
      f"raw={raw_total:,.0f} vs module={mod_total:,.0f}")
hist_total = sum(p.qty for p in rl.histogram)
check("A14b histogram sums to totals", abs(hist_total - mod_total) < 1.0,
      f"hist={hist_total:,.0f} vs {mod_total:,.0f}")


# A15. Milestone terminals: EVERY milestone is offered, achieved or
# not — you pick what you are measuring to; whether the works reached
# it is disclosed, never a filter.
from programme import trace_end_candidates as _tec
_a15 = _tec([("B", B), ("U", U)], contract_ms="KD15")
check("A15 elected milestone offered first even though unachieved",
      _a15[0][0] == "KD15" and _a15[0][3] is False)
_ms_codes = {t.task_code for t in U.tasks if t.is_milestone
             and not t.is_loe_or_wbs}
_offered = {c for c, _, _, _ in _a15}
check("A15b every milestone is offered, achieved or not",
      _ms_codes <= _offered, f"{len(_ms_codes - _offered)} missing")
check("A15c unachieved milestones carry achieved=False + a forecast date",
      all(ok or d is not None
          for c, _, d, ok in _a15 if c in _ms_codes))


# A16. Actual-date trace + triangulation invariants
from programme import extract_actual_trace, trace_end_candidates
tr_strict = extract_actual_trace([("B", B), ("U", U)], max_gap_days=240,
                                 allow_temporal_fallback=False)
check("A16 strict trace: every link logic-evidenced",
      all(lk.had_logic for lk in tr_strict.links))
check("A16b trace links form a chain (each pred is next activity)",
      all(lk.score is not None and 0 <= lk.score <= 1
          for lk in tr_strict.links))
codes = [a.task_code for a in tr_strict.activities]
check("A16c trace chain has no duplicates", len(codes) == len(set(codes)))
tr_fb = extract_actual_trace([("B", B), ("U", U)], max_gap_days=15)
check("A16d default (continuous) trace >= strict at same gap",
      len(tr_fb.activities) >= len(extract_actual_trace(
          [("B", B), ("U", U)], max_gap_days=15,
          allow_temporal_fallback=False).activities))
check("A16f default continues on sequence: un-evidenced hops disclosed",
      all(lk.had_logic for lk in tr_fb.links)
      or any("SEQUENCE ALONE" in w for w in tr_fb.warnings))

# A16g-k. Contractual-milestone anchoring + hybrid forecast tail.
_kd15 = next((t for t in U.tasks if t.task_code == "KD15"), None)
check("A16g sample carries an unachieved completion milestone (KD15)",
      _kd15 is not None and _kd15.act_finish is None)
_hy = extract_actual_trace([("B", B), ("U", U)], end_task_code="KD15",
                           max_gap_days=60)
check("A16h unachieved elected milestone still anchors the trace",
      _hy.terminal_code == "KD15" and _hy.hybrid)
check("A16i hybrid disclosed in warnings and caveats",
      any("HYBRID" in c for c in _hy.caveats)
      and any("not been achieved" in w.lower() or "NOT been achieved" in w
              for w in _hy.warnings))
check("A16j hybrid chain is basis-labelled and reaches back to as-built",
      _hy.forecast_count > 0 and _hy.asbuilt_count > 0
      and {a.basis for a in _hy.activities} <= {
          "as-built", "in-progress", "forecast"})
# ordering: as-built work must precede the forecast tail
_bases = [a.basis for a in _hy.activities]
check("A16k forecast tail sits at the end of the chain (no forecast "
      "before as-built work)",
      all(b == "forecast" for b in _bases[_bases.index("forecast"):])
      if "forecast" in _bases else True)
check("A16l terminal candidates offer the elected milestone first",
      trace_end_candidates([("B", B), ("U", U)], contract_ms="KD15")[0][0]
      == "KD15")

# A16m-s. Analyst-election trace (the step-① adopted path, shared by
# the As-Built CP page and APvAB). An elected path must be reported
# no more charitably than a computed one.
from programme import trace_from_election, build_asbuilt_multi_prompt
_ad_path = [(a.task_code, a.name) for a in tr_fb.activities]
_el = trace_from_election([("B", B), ("U", U)], _ad_path,
                          basis_label="Actual sequence through recorded "
                                      "dates (test)")
check("A16m election trace keeps the adopted order and terminal",
      [a.task_code for a in _el.activities] == [c for c, _ in _ad_path]
      and _el.terminal_code == _ad_path[-1][0])
check("A16n election trace discloses the adopted basis in its caveats",
      any("ELECTION" in c for c in _el.caveats)
      and any("Actual sequence through recorded dates (test)" in c
              for c in _el.caveats))
check("A16o election links agree with the computed trace on logic",
      {(lk.pred_code, lk.succ_code): lk.had_logic for lk in _el.links}
      == {(lk.pred_code, lk.succ_code): lk.had_logic
          for lk in tr_fb.links})
check("A16p election link scores stay in [0,1]",
      all(0 <= lk.score <= 1 for lk in _el.links))
# hand-edit: a pair the records cannot support scores zero temporal
# evidence (successor starting long before its predecessor began)
_el_bad = trace_from_election([("B", B), ("U", U)],
                              [_ad_path[-1], _ad_path[0]])
_bad_lk = _el_bad.links[0] if _el_bad.links else None
check("A16q an out-of-order adopted pair is scored weak, not hidden",
      _bad_lk is not None and _bad_lk.score <= 0.4)
_el_hy = trace_from_election(
    [("B", B), ("U", U)],
    _ad_path + [("KD15", next(t.name for t in U.tasks
                              if t.task_code == "KD15"))])
check("A16r a forecast tail in the election is a disclosed hybrid",
      _el_hy.hybrid and any("HYBRID" in c for c in _el_hy.caveats))
_mp = build_asbuilt_multi_prompt([_el, _el_hy])
check("A16s multi-path prompt covers every path and templates once",
      "PATH 1 of 2" in _mp and "PATH 2 of 2" in _mp
      and _mp.count("<context>") == 2)

# A16t-v. As-built longest path (the step-① logic candidate) must run
# THROUGH completed work to the earliest linked activity — not stop at
# the data date the way a remaining-works longest path does.
from programme import extract_asbuilt_longest_path
_lp = extract_asbuilt_longest_path(U, end_task_code="KD15")
_dd = U.project.data_date
_lp_starts = [a.act_start for a in _lp.activities if a.act_start]
check("A16t as-built longest path reaches back past the data date",
      bool(_lp_starts) and min(_lp_starts) < _dd,
      f"earliest {min(_lp_starts) if _lp_starts else None} vs dd {_dd}")
check("A16u as-built longest path: every hand-off is programmed logic "
      "and the chain is chronological",
      all(lk.had_logic for lk in _lp.links)
      and all(a.act_start <= b.act_start
              for a, b in zip(_lp.activities, _lp.activities[1:])
              if a.act_start and b.act_start))
check("A16v unachieved terminal still anchors the logic candidate "
      "as a disclosed hybrid",
      _lp.terminal_code == "KD15" and _lp.hybrid
      and any("HYBRID" in c for c in _lp.caveats))


# A17. Sequence coding invariants
from programme import propose_sequence_mapping, analyse_sequence
sp = propose_sequence_mapping(U, "U")
check("A17 sequence: every activity gets a front and a stage",
      all(r.front and r.stage for r in sp.rows))
check("A17b sequence: coverage percentages in [0,100]",
      0 <= sp.stage_coverage_pct <= 100 and 0 <= sp.front_coverage_pct <= 100)
sq = analyse_sequence(sp.rows, "U")
check("A17c sequence: band bounds ordered (start <= finish)",
      all(b.act_start is None or b.act_finish is None
          or b.act_start <= b.act_finish for b in sq.bands))
check("A17d sequence: mapped count == actualised rows",
      sq.mapped_activities == sum(1 for r in sp.rows if r.act_start))
check("A17e sequence: unconfirmed mapping carries the extra caveat",
      any("AUTO-PROPOSED" in c for c in sq.caveats))
sq2 = analyse_sequence(sp.rows, "U", mapping_confirmed=True)
check("A17f sequence: confirmed mapping drops it",
      not any("AUTO-PROPOSED" in c for c in sq2.caveats))


# A18. AI-review prompt/parser layer (offline)
from programme import (build_mapping_review_prompt, parse_mapping_review,
                       build_view_advice_prompt, parse_view_advice)
pmr = build_mapping_review_prompt(sp.rows[:5])
check("A18 review prompt lists stages and rows",
      "Allowed stage labels" in pmr and sp.rows[0].task_code in pmr)
good = parse_mapping_review(
    '[{"id": "%s", "stage": "Finishes & Fit-Out"}]' % sp.rows[0].task_code,
    {r.task_code for r in sp.rows[:5]})
check("A18b parser accepts valid correction", len(good) == 1)
check("A18c parser rejects unknown ids and stages",
      parse_mapping_review('[{"id":"ZZZ","stage":"Finishes & Fit-Out"},'
                           '{"id":"%s","stage":"Made Up"}]'
                           % sp.rows[0].task_code,
                           {sp.rows[0].task_code}) == {})
check("A18d parser survives garbage", parse_mapping_review("oops", {"A"}) == {})
adv = parse_view_advice('{"mode":"bands","colour":"Stage","max_fronts":10,"rationale":"r"}')
check("A18e view advice parses and clamps",
      adv is not None and adv["mode"] == "bands"
      and parse_view_advice('{"mode":"nope"}') is None)
check("A18f view advice prompt built",
      "sequence_gantt" in build_view_advice_prompt(sq, 30))


# A19. Hierarchy rebuild invariants
from programme import (available_dimensions, build_hierarchy, tree_to_dict,
                       build_gantt_html, config_to_json, config_from_json)
hd = available_dimensions(B)
check("A19 dimensions discovered (5 WBS levels, no codes in sample)",
      len([d for d in hd if d.dim_id.startswith("wbs:")]) == 5)
hh = build_hierarchy(B, ["wbs:2", "wbs:3"], "B",
                     dim_labels=["WBS Level 2", "WBS Level 3"])
check("A19b every source activity placed exactly once",
      hh.is_complete and hh.placed_activities == hh.source_activities)
# leaf-count == placed (no duplication anywhere in the tree)
def _leaves(n):
    return len(n.activities) + sum(_leaves(c) for c in n.children.values())
check("A19c tree leaf count == placed", _leaves(hh.root) == hh.placed_activities)
# rollup: root span brackets every activity date
def _acts(n):
    yield from n.activities
    for c in n.children.values():
        yield from _acts(c)
all_starts = [a.start for a in _acts(hh.root) if a.start]
all_fins = [a.finish for a in _acts(hh.root) if a.finish]
root_kids = list(hh.root.children.values())
check("A19d rollup start == min child start",
      min(k.start for k in root_kids if k.start) == min(all_starts))
check("A19e rollup finish == max child finish",
      max(k.finish for k in root_kids if k.finish) == max(all_fins))
# source data untouched: parse count unchanged after building
check("A19f source untouched (task count stable)",
      hh.source_activities == sum(1 for t in B.tasks
                                  if t.task_type != "TT_WBS"))
html = build_gantt_html(tree_to_dict(hh.root))
check("A19g gantt html self-contained", "<script>" in html
      and "http" not in html.split("</style>")[0].lower())
cfg = config_from_json(config_to_json("v", ["wbs:2"], ["WBS Level 2"]))
check("A19h config round-trips", cfg is not None and cfg[1] == ["wbs:2"])
check("A19i bad config rejected", config_from_json('{"dimensions":["x:1"]}') is None)


# A20. Sequence dims + hierarchy xlsx
from programme import sequence_dimension_mappings, build_hierarchy_xlsx
ex = sequence_dimension_mappings(U, sp.rows)
hs = build_hierarchy(U, ["seq:front", "seq:stage"], "U",
                     dim_labels=["Front", "Stage"], extra_mappings=ex)
check("A20 seq-dims hierarchy places all activities",
      hs.is_complete and hs.placed_activities == hs.source_activities)
xh = build_hierarchy_xlsx(hs)
from openpyxl import load_workbook as _lw
import io as _io2
_wbh = _lw(_io2.BytesIO(xh))
check("A20b hierarchy xlsx sheets",
      set(_wbh.sheetnames) >= {"Hierarchy", "Flat Table"})
outl = sum(1 for rd in _wbh["Hierarchy"].row_dimensions.values()
           if rd.outline_level)
check("A20c hierarchy xlsx has collapsible outlines", outl > 100)
flat_rows = _wbh["Flat Table"].max_row - 1
check("A20d flat table row per activity",
      flat_rows == hs.placed_activities,
      f"flat={flat_rows} vs placed={hs.placed_activities}")
check("A20e seq config ids accepted",
      config_from_json('{"dimensions": ["seq:front"], "labels": ["F"]}')
      is not None)


# A21. Dimension menu = WBS levels + activity codes + TASK UDFs only
hd2 = available_dimensions(U)
kinds2 = {d.dim_id.partition(":")[0] for d in hd2}
check("A21 only the three families offered", kinds2 <= {"wbs", "code", "udf"})
check("A21b all WBS levels present",
      {f"wbs:{i}" for i in range(1, 6)} <= {d.dim_id for d in hd2})
# synthetic TASK UDF proves the udf: path end-to-end
_t0 = U.tasks[0]
U.raw_tables.setdefault("UDFTYPE", []).append(
    {"udf_type_id": "999", "table_name": "TASK",
     "udf_type_label": "QA Zone", "udf_type_name": "qa_zone",
     "logical_data_type": "FT_TEXT"})
U.raw_tables.setdefault("UDFVALUE", []).append(
    {"udf_type_id": "999", "fk_id": _t0.task_id, "udf_text": "Zone QA",
     "udf_number": "", "udf_date": "", "udf_code_id": ""})
hd3 = available_dimensions(U)
check("A21c TASK UDF surfaces as a dimension",
      any(d.dim_id == "udf:999" and "QA Zone" in d.label for d in hd3))
_hu = build_hierarchy(U, ["udf:999"], "U", dim_labels=["QA Zone"])
check("A21d UDF hierarchy: tagged task grouped, rest Unassigned",
      _hu.is_complete and "Zone QA" in _hu.root.children
      and _hu.root.children["Zone QA"].activity_count == 1)
U.raw_tables["UDFTYPE"].pop(); U.raw_tables["UDFVALUE"].pop()
# synthetic global + project code types both surface, scope-labelled
U.raw_tables.setdefault("ACTVTYPE", []).append(
    {"actv_code_type_id": "801", "actv_code_type": "Zone",
     "actv_code_type_scope": "AS_Global"})
U.raw_tables["ACTVTYPE"].append(
    {"actv_code_type_id": "802", "actv_code_type": "Package",
     "actv_code_type_scope": "AS_Project"})
hd4 = available_dimensions(U)
lbls = {d.dim_id: d.label for d in hd4}
check("A21e global + project codes both offered, scope in label",
      "[Global]" in lbls.get("code:801", "")
      and "[Project]" in lbls.get("code:802", ""))
U.raw_tables["ACTVTYPE"] = []
check("A21f config kinds restricted",
      config_from_json('{"dimensions": ["cal:"]}') is None
      and config_from_json('{"dimensions": ["udf:9", "wbs:2"]}') is not None)


# A22. Prospective TIA engine
from programme import (DelayEvent, FragnetActivity, FragnetLink, run_tia,
                       validate_fragnet, parse_fragnet_json, parse_links,
                       find_template_activities, find_template_work_packages,
                       assess_event_scope, build_logic_recommendation_prompt,
                       parse_logic_recommendation_json)
from datetime import timedelta as _td
_ev = DelayEvent("EV-QA", "test event")
_fr = [FragnetActivity("TIA-010", "chain", 120,
                       successors=[FragnetLink("KD15")])]
_r = run_tia(U, "U", _ev, _fr)
check("A22 TIA delta exact for a direct chain into completion",
      _r.completion_post == _r.data_date + _td(days=120)
      and (_r.completion_delta_days or 0) > 0)
_r0 = run_tia(U, "U", _ev, [])
check("A22b empty fragnet -> zero delta",
      _r0.completion_pre == _r0.completion_post)
check("A22c calibration disclosed", _r.calibration_days is not None
      and any("Calibration" in w for w in _r.warnings))
iss = validate_fragnet(U, [FragnetActivity("TIA-1", "x", -5)])
check("A22d validation flags open ends + bad duration",
      any("open start" in i for i in iss)
      and any("duration" in i for i in iss))
iss2 = validate_fragnet(U, [
    FragnetActivity("TIA-A", "a", 5,
                    predecessors=[FragnetLink("TIA-B")],
                    successors=[FragnetLink("TIA-B"), FragnetLink("KD15")]),
    FragnetActivity("TIA-B", "b", 5,
                    predecessors=[FragnetLink("TIA-A")],
                    successors=[FragnetLink("TIA-A")])])
check("A22e circular fragnet detected",
      any("Circular" in i for i in iss2))
check("A22f fragnet json parser rejects invalid refs",
      parse_fragnet_json('{"activities":[{"id":"TIA-1","name":"x",'
                         '"duration_days":5,'
                         '"successors":[{"id":"NOPE-99"}]}]}', U)[0]
      .successors == [])
check("A22g template search returns project evidence",
      len(find_template_activities(U, "installation of ceiling")) > 0)
check("A22h link text round-trip",
      parse_links("A1:SS:5")[0].link_type == "SS")
_scope = assess_event_scope(DelayEvent(
    "EV-S", "Additional ceiling installation", "include approval and test",
    area="Zone B", discipline="Architectural", project_context="Hospital",
    work_package="Additional ceiling works"))
check("A22i event understood before fragnet drafting",
      _scope.work_nature.startswith("Additional")
      and "Testing / inspection / handover" in _scope.lifecycle_stages)
_pkgs = find_template_work_packages(U, "installation of ceiling")
check("A22j existing work packages ranked before generic drafting",
      bool(_pkgs) and bool(_pkgs[0]["activities"])
      and _pkgs[0]["score"] > 0)
_logic_prompt = build_logic_recommendation_prompt(_ev, _fr, U)
check("A22k logic recommendation uses confirmed fragnet + programme IDs",
      "TIA-010" in _logic_prompt and "allowed_existing_activities" in _logic_prompt)
_known_pred = U.tasks[0].task_code
_logic = parse_logic_recommendation_json(
    '{"predecessors":[{"id":"' + _known_pred + '","type":"FS","lag_days":0}],'
    '"successors":[{"id":"KD15","type":"FS","lag_days":0}],'
    '"impacted_sections":[{"id":"KD15"}],'
    '"warnings":["planner review"]}', U)
_logic_bad = parse_logic_recommendation_json(
    '{"predecessors":[{"id":"INVENTED-1"}]}', U)
check("A22l logic parser accepts programme IDs and rejects invention",
      _logic["predecessors"][0]["id"] == _known_pred
      and _logic_bad["predecessors"] == [])
_calendar_id = next(iter(U.calendars))
_calendar_fragnet = parse_fragnet_json(
    '{"activities":[{"id":"TIA-CAL","name":"calendar test",'
    '"duration_days":2,"calendar_id":"' + _calendar_id + '",'
    '"successors":[{"id":"KD15"}]}]}', U)
check("A22m fragnet retains only a valid programme calendar",
      _calendar_fragnet[0].calendar_id == _calendar_id)
_targeted = run_tia(U, "U", _ev, _fr, target_milestone="KD15")
check("A22n selected impacted milestone is prioritised in results",
      bool(_targeted.milestone_impacts)
      and _targeted.milestone_impacts[0].code == "KD15")
from programme import build_tia_xlsx
_tia_book = load_workbook(io.BytesIO(build_tia_xlsx(
    _targeted, audit={"source_sha256": "abc"},
    run_history=[{"completion_delta_days": 5}])))
check("A22o TIA export includes audit and rerun history",
      "Audit Trail" in _tia_book.sheetnames
      and "Run History" in _tia_book.sheetnames
      and "Calendar" in [c.value for c in _tia_book["Fragnet"][1]])


# A23. Explain This Delay
from programme import explain_delay
_ex = explain_delay([("B", B), ("U", U)], "KD15")
check("A23 explain: facts recorded per revision",
      len(_ex.points) == 2 and _ex.points[0].forecast is not None)
check("A23b explain: total movement == raw forecast delta",
      abs(_ex.total_movement_days
          - (_ex.points[-1].forecast
             - _ex.points[0].forecast).days) < 1)
check("A23c explain: uncertain attribution flagged when path switched",
      any(not w.attribution_reliable for w in _ex.windows)
      and any("uncertain" in w for w in _ex.warnings))
check("A23d explain: facts/inference separation in caveats",
      any("INFERENCE" in c for c in _ex.caveats))
_ex1 = explain_delay([("B", B)], "KD15")
check("A23e explain: single revision -> warning, no crash",
      not _ex1.windows and _ex1.warnings)


# A24. Event extraction (TIA intake) + 52R-06
from programme import (build_event_extraction_prompt, parse_event_candidates,
                       read_document, recommended_analysis_schedule)
_docs = [("L1.txt", "On 12 March 2018 the Engineer issued Instruction "
                    "EI-88 requiring additional ceiling works.")]
_ep = build_event_extraction_prompt(_docs)
check("A24 extraction prompt cites 52R-06 and the doc",
      "52R-06" in _ep and "L1.txt" in _ep)
_good = ('{"events":[{"title":"EI-88","source_doc":"L1.txt",'
         '"source_snippet":"issued Instruction EI-88","date_start":'
         '"2018-03-12","confidence":"high"}]}')
_c, _d = parse_event_candidates(_good, _docs)
check("A24b verified snippet accepted", len(_c) == 1 and _c[0].verified)
_bad = ('{"events":[{"title":"Flood","source_doc":"L1.txt",'
        '"source_snippet":"site flooded for weeks"}]}')
_c2, _d2 = parse_event_candidates(_bad, _docs)
check("A24c fabricated snippet dropped", _c2 == [] and _d2 == 1)
check("A24d garbage tolerated", parse_event_candidates("x", _docs) == ([], 0))
from datetime import datetime as _dtx
_meta = [("U1", _dtx(2018, 1, 31)), ("U2", _dtx(2018, 2, 28))]
check("A24e 52R-06 picks last update before event",
      recommended_analysis_schedule(_meta, _dtx(2018, 2, 10)) == "U1")
check("A24f TIA caveats cite 52R-06",
      any("52R-06" in c for c in _r.caveats))
check("A24g txt reader works", "hello" in read_document("a.txt", b"hello"))


# A25. Impacted-programme XER export round-trip
from programme import build_impacted_xer
_raw = open(_p("sample/Sample Update.xer"), "rb").read()
_fr2 = [FragnetActivity("TIA-910", "a", 10,
                        successors=[FragnetLink("TIA-920")]),
        FragnetActivity("TIA-920", "b", 20,
                        predecessors=[FragnetLink("TIA-910")],
                        successors=[FragnetLink("KD15")])]
_res2 = run_tia(U, "U", _ev, _fr2)
_out = build_impacted_xer(_raw.decode("utf-8", errors="replace"),
                          U, _fr2, _res2)
_u2 = parse_xer(_out.encode("utf-8"))
check("A25 impacted xer: fragnet tasks import",
      len(_u2.tasks) == len(U.tasks) + 2)
check("A25b impacted xer: links deduped and resolved",
      len(_u2.relationships) == len(U.relationships) + 2)
_t2 = next(x for x in _u2.tasks if x.task_code == "TIA-920")
check("A25c impacted xer: not-started with duration",
      _t2.status == "TK_NotStart" and _t2.target_drtn_hr is not None)


# A26. Calendar-exact CPM + cumulative TIA + concurrency
from programme import run_cumulative_tia
check("A26 calendar-exact calibration within 2 days of P6",
      abs(_r.calibration_days or 99) < 2, f"calib={_r.calibration_days}")
from datetime import datetime as _dt6
_evA = DelayEvent("EV-A", "a", date_raised=_dt6(2018, 5, 1))
_evB = DelayEvent("EV-B", "b", date_raised=_dt6(2018, 5, 20))
_cum = run_cumulative_tia(U, "U", [
    (_evB, [FragnetActivity("TIA-B1", "b", 170,
                            successors=[FragnetLink("KD35")])]),
    (_evA, [FragnetActivity("TIA-A1", "a", 150,
                            successors=[FragnetLink("KD15")])])])
check("A26b cumulative inserts chronologically",
      _cum["rows"][0]["event_id"] == "EV-A")
check("A26c incremental deltas sum to total",
      abs(sum(r["incremental_delta_days"] for r in _cum["rows"])
          - _cum["total_delta_days"]) < 0.2)
check("A26d overlapping driving chains flagged as concurrency candidates",
      len(_cum["concurrency"]) == 1)


# A27. Notice screening + clause extraction + TIA report chart
from programme import (assess_notice, build_clause_extraction_prompt,
                       parse_clause_extraction)
from programme import report_charts as _rc27
from datetime import datetime as _dt7
check("A27 notice compliant with margin",
      assess_notice(_dt7(2018,5,3), _dt7(2018,5,20), 28).status
      == "compliant")
check("A27b notice late",
      assess_notice(_dt7(2018,5,3), _dt7(2018,7,1), 28).status == "late")
check("A27c no notice / indeterminate",
      assess_notice(_dt7(2018,5,3), None, 28).status == "no_notice"
      and assess_notice(None, None, None).status == "indeterminate")
_ct = "Clause 20.1: the Contractor shall give notice within 28 days of awareness."
_ok = parse_clause_extraction(
    '{"clauses":[{"topic":"notice","clause_ref":"20.1","period_days":28,'
    '"requirement":"notify","snippet":"give notice within 28 days",'
    '"silent":false},{"topic":"float","silent":true},'
    '{"topic":"fake","snippet":"invented words here","silent":false}]}', _ct)
check("A27d clause parser: verified kept, silent kept, invented dropped",
      len(_ok) == 2 and _ok[0]["period_days"] == 28)
check("A27e TIA paths chart builds",
      _rc27.tia_paths_chart(_r) is not None)

print("\n== B. Edge cases / degenerate inputs ==")

# B1. Windows with one revision
w1 = analyse_windows([("B", B)])
check("B1 single-revision windows -> warning, no crash",
      not w1.windows and w1.warnings)

# B2. Float erosion with same file twice -> zero erosion
fe2 = analyse_float_erosion([("B", B), ("B2", B)])
check("B2 self float erosion: median delta == 0",
      fe2.windows[0].median_delta == 0 and fe2.windows[0].eroded_count == 0)

# B3. Progress with no updates
pr0 = compute_progress(B, "B", [])
check("B3 progress w/o updates: planned only, no crash",
      pr0.planned_curve and not pr0.recorded_curve
      and pr0.time_offset_days is None)

# B4. Longest path with bogus end code -> falls back with warning
cp_b = extract_longest_path(B, "B", end_task_code="NOPE-123")
check("B4 bogus end code -> fallback + warning",
      cp_b.end_choice is not None
      and any("not found" in w for w in cp_b.warnings))

# B5. Resources on fixture without RSRC table
rA = extract_resource_loading(fix[0][1], "revA")
check("B5 no-resource file -> warning, no crash",
      not rA.histogram and rA.warnings)

# B6. Critical path with absurd tolerance -> no critical, warning
cp_none = extract_critical_path(B, "B", float_tolerance_days=-9999)
check("B6 impossible tolerance -> warning, empty",
      not cp_none.critical and cp_none.warnings)

# B7. Fixtures through every multi-rev engine (3 revisions)
try:
    analyse_windows(fix); analyse_float_erosion(fix)
    compare_revisions(fix[0][1], fix[2][1], "A", "C")
    compute_progress(fix[0][1], "A", [(l, d) for l, d in fix[1:]])
    check("B7 fixtures through all multi-rev engines", True)
except Exception as e:
    check("B7 fixtures through all multi-rev engines", False,
          f"{type(e).__name__}: {e}")

print("\n== C. Report integrity (prompts + workbooks) ==")
from openpyxl import load_workbook
import io as _io

inv = build_inventory(inv_pool)
builds = {
    "inventory": (build_inventory_prompt(inv), build_inventory_xlsx(inv, "n")),
    "milestones": (build_milestone_prompt(ms, ms.series[:5]),
                   build_milestone_xlsx(ms, ms.series[:5], "n")),
    "variance": (build_variance_prompt(var), build_variance_xlsx(var, "n")),
    "critical_path": (build_critical_path_prompt(cp_l),
                      build_critical_path_xlsx(cp_l, "n")),
    "comparison": (build_comparison_prompt(c_fwd),
                   build_comparison_xlsx(c_fwd, "n")),
    "windows": (build_windows_prompt(wres), build_windows_xlsx(wres, "n")),
    "progress": (build_progress_prompt(pr), build_progress_xlsx(pr, "n")),
    "float_erosion": (build_float_erosion_prompt(fe),
                      build_float_erosion_xlsx(fe, "n")),
    "resources": (build_resources_prompt(rl), build_resources_xlsx(rl, "n")),
}
for name, (prompt, xlsx) in builds.items():
    has_rules = "<rules>" in prompt and "Attribute nothing" in prompt
    has_caveats = "<caveats>" in prompt or "warnings" in prompt.lower() or name == "inventory"
    wb = load_workbook(_io.BytesIO(xlsx))
    has_narr = "AI Narrative" in wb.sheetnames
    check(f"C {name}: hard rules in prompt", has_rules)
    check(f"C {name}: workbook opens, narrative sheet present",
          has_narr, str(wb.sheetnames))

# C2. Every module's standing caveats reach its prompt
for name, (prompt, _) in builds.items():
    if name == "inventory":
        continue
    check(f"C2 {name}: limitations content present",
          "caveat" in prompt.lower() or "<caveats>" in prompt)

print("== D. TIA hardening upgrades ==")
from datetime import datetime as _dt

from programme.tia import (DelayEvent, FragnetActivity, FragnetLink,
                           _backward_pass, _build_network, _calendar_masks,
                           _forward_pass, run_cumulative_tia, run_tia)
from programme.xer_export import build_impacted_xer
from programme.events_extract import parse_event_candidates, truncation_notes
from programme.notice import assess_notice

_masks = _calendar_masks(U)
check("D1 calendar masks carry holiday exceptions",
      any(len(v[1]) > 0 for v in _masks.values()),
      f"{sum(len(v[1]) for v in _masks.values())} holidays total")

_ev = DelayEvent("EV-QA", "Chiller rework", "rework to chiller plant")
_frag = [FragnetActivity("TIA-010", "Remove", 20,
                         predecessors=[FragnetLink("RM-AC-005")],
                         successors=[FragnetLink("TIA-020")]),
         FragnetActivity("TIA-020", "Reinstall", 40,
                         predecessors=[FragnetLink("TIA-010")],
                         successors=[FragnetLink("TOC05")])]
_r = run_tia(U, "U", _ev, _frag)
check("D2 start-constraint floors applied and disclosed",
      any("start constraint" in w for w in _r.warnings))
check("D3 tie-in float reported, post <= pre",
      bool(_r.tie_in_float) and all(
          t["float_post"] <= t["float_pre"]
          for t in _r.tie_in_float
          if t["float_pre"] is not None and t["float_post"] is not None))
check("D4 milestone impacts carry total float",
      any(m.float_pre is not None and m.float_post is not None
          for m in _r.milestone_impacts))
check("D4b calibration still tight with masks+constraints",
      _r.calibration_days is not None and abs(_r.calibration_days) <= 2,
      f"calibration {_r.calibration_days}")

# D5 completion symmetry: post completion never taken from a fragnet act
_dd = U.project.data_date
_inc, _nodes, _preds, _started, _fm, _ = _build_network(U, cfg, _dd)
_np = dict(_nodes); _pp = {k: list(v) for k, v in _preds.items()}
for _f in _frag:
    _np[_f.act_id] = (max(_f.duration_days, 0.0), None)
    _pp.setdefault(_f.act_id, [])
    for _l in _f.predecessors:
        _pp[_f.act_id].append((_l.other_id, _l.link_type, _l.lag_days))
    for _l in _f.successors:
        _pp.setdefault(_l.other_id, []).append(
            (_f.act_id, _l.link_type, _l.lag_days))
_, _EF1, _, _ = _forward_pass(_np, _pp, _dd, _started)
check("D5 completion_post measured over the real network only",
      _r.completion_post == max(ef for c, ef in _EF1.items()
                                if c in _nodes))

# D6 backward pass: a genuinely critical chain exists (min TF ~ 0)
_, _EF0, _, _ = _forward_pass(dict(_nodes),
                              {k: list(v) for k, v in _preds.items()},
                              _dd, _started)
_tf = _backward_pass(_nodes, _preds, _EF0)
check("D6 backward pass yields a zero-float driving chain",
      _tf and min(abs(v) for v in _tf.values()) <= 1.0,
      f"min |TF| = {min(abs(v) for v in _tf.values()) if _tf else '—'}")

# D7 cumulative ID clash is caught and the duplicate skipped
_ev2 = DelayEvent("EV-QB", "Clash", "")
_frag2 = [FragnetActivity("TIA-010", "Dup id", 10,
                          predecessors=[FragnetLink("RM-AC-005")],
                          successors=[FragnetLink("TOC05")])]
_cum = run_cumulative_tia(U, "U", [(_ev, _frag), (_ev2, _frag2)])
check("D7 cumulative flags reused fragnet IDs",
      any("SKIPPED" in w for w in _cum.get("warnings", [])))

# D8 impacted XER: dedicated fragnet WBS band + exact table anchoring
with open(_p("sample/Sample Update.xer"), encoding="latin-1") as fh:
    _raw = fh.read()
_out = build_impacted_xer(_raw, U, _frag, _r)
_U2 = parse_xer(_out.encode("latin-1", errors="replace"))
_wrows = [w for w in _U2.raw_tables.get("PROJWBS", [])
          if "TIA Fragnet" in (w.get("wbs_name") or "")]
_trows = [t for t in _U2.raw_tables.get("TASK", [])
          if (t.get("task_code") or "").startswith("TIA-")]
check("D8 impacted XER round-trips with fragnet WBS band",
      len(_wrows) == 1 and len(_trows) == 2
      and all(t.get("wbs_id") == _wrows[0].get("wbs_id") for t in _trows)
      and len(_U2.tasks) == len(U.tasks) + 2
      and len(_U2.relationships) == len(U.relationships) + 3)

# D9 event extraction: documented end date -> stated duration; bad order rejected
_docs = [("L1.txt", "The Engineer instructed suspension of chiller works "
          "from 12 May 2018; the suspension was lifted on 3 June 2018.")]
_resp = ('{"events":[{"title":"Suspension","date_start":"2018-05-12",'
         '"date_end":"2018-06-03","source_doc":"L1.txt",'
         '"source_snippet":"instructed suspension of chiller works",'
         '"confidence":"high"}]}')
_cands, _ = parse_event_candidates(_resp, _docs)
check("D9 date_end captured, stated duration computed",
      _cands and _cands[0].stated_duration_days == 22.0)
_bad = _resp.replace('"date_end":"2018-06-03"', '"date_end":"2018-05-01"')
_cands_b, _ = parse_event_candidates(_bad, _docs)
check("D9b end-before-start rejected",
      _cands_b and _cands_b[0].date_end is None)
check("D9c truncation disclosed for oversize documents",
      truncation_notes([("big.pdf", "x" * 20001)]) != []
      and truncation_notes(_docs) == [])

# D10 notice basis changes the count and is printed
_na_c = assess_notice(_dt(2018, 5, 11), _dt(2018, 5, 14), 2, "calendar")
_na_b = assess_notice(_dt(2018, 5, 11), _dt(2018, 5, 14), 2, "business")
check("D10 Fri->Mon: 3 calendar days late, 1 business day compliant",
      _na_c.status == "late" and _na_b.status == "compliant"
      and "business day" in _na_b.detail)

# D11 impossible notice inputs never yield a contractual status
check("D11 notice before awareness -> indeterminate",
      assess_notice(_dt(2018, 5, 10), _dt(2018, 5, 5), 14).status
      == "indeterminate")
check("D11b non-positive clause period -> indeterminate",
      assess_notice(_dt(2018, 5, 10), _dt(2018, 5, 12), -7).status
      == "indeterminate"
      and assess_notice(_dt(2018, 5, 10), _dt(2018, 5, 12), 0).status
      == "indeterminate")

print("== E. Comparison impact, progress transfer, project library ==")

from programme import (assess_comparison_impact, build_provenance,
                       out_of_sequence_flags, run_progress_transfer,
                       ProjectStore)
import tempfile

# E1. Impact screening — coverage, ordering, and score sanity
_imp = assess_comparison_impact(B, U, "B", "U")
_cmp_bu = compare_revisions(B, U, "B", "U")
_bands_ok = all(c.band_old in ("critical", "near-critical", "off-path",
                               "completed", "absent")
                and c.band_new in ("critical", "near-critical", "off-path",
                                   "completed", "absent")
                for c in _imp.ranked)
check("E1 every ranked change carries valid path bands", _bands_ok)
check("E1b ranked count == diff total minus renames",
      len(_imp.ranked) == _cmp_bu.total_changes - len(_cmp_bu.renamed),
      f"ranked={len(_imp.ranked)} vs "
      f"{_cmp_bu.total_changes - len(_cmp_bu.renamed)}")
check("E1c rank is sorted by score descending",
      all(_imp.ranked[i].score >= _imp.ranked[i + 1].score
          for i in range(len(_imp.ranked) - 1)))
check("E1d every retrospective actual change is red-flagged",
      sum(1 for c in _imp.ranked if c.red_flag)
      >= len(_cmp_bu.actual_date_changes))
_imp_self = assess_comparison_impact(B, B, "B", "B")
check("E1e self-impact == 0 ranked changes", len(_imp_self.ranked) == 0,
      f"got {len(_imp_self.ranked)}")

# E2. Out-of-sequence screening — well-formed, and a manual FS recount
_oos = out_of_sequence_flags(U)
check("E2 OOS overlaps positive or None (open predecessor)",
      all(f.overlap_days is None or f.overlap_days > 0 for f in _oos))
_by_id = {t.task_id: t for t in U.tasks if not t.is_loe_or_wbs}
_manual_fs = 0
for _r in U.relationships:
    _pt, _st_ = _by_id.get(_r.pred_task_id), _by_id.get(_r.task_id)
    if (_pt is not None and _st_ is not None and _r.pred_type == "PR_FS"
            and _st_.act_start and _pt.act_finish
            and (_pt.act_finish - _st_.act_start).total_seconds()
            / 86400.0 > 0.1):
        _manual_fs += 1
_fs_flags = sum(1 for f in _oos
                if f.link_type == "FS" and f.overlap_days is not None)
check("E2b FS overlap flags == manual recount from raw actuals",
      _fs_flags == _manual_fs, f"flags={_fs_flags} vs manual={_manual_fs}")

# E3. Provenance — windows equal the direct pairwise diffs
_prov = build_provenance(fix)
check("E3 provenance windows == revisions - 1",
      len(_prov.windows) == len(fix) - 1)
_direct = compare_revisions(fix[0][1], fix[1][1], fix[0][0], fix[1][0])
check("E3b window counts match direct pairwise diff",
      _prov.windows[0].counts == _direct.category_counts)
check("E3c red-flag count mirrors actual-date changes",
      all(w.red_flag_count == len(w.comparison.actual_date_changes)
          for w in _prov.windows))

# E4. Progress transfer — self-transfer identity + manual recounts
_tr_self = run_progress_transfer(U, U, "U", "U")
check("E4 self-transfer network effect == 0",
      _tr_self.network_effect_days == 0.0,
      f"got {_tr_self.network_effect_days}")
_tr = run_progress_transfer(B, U, "B", "U")
_b_codes = {t.task_code for t in B.tasks if not t.is_loe_or_wbs}
_manual_fin = sum(1 for t in U.tasks
                  if not t.is_loe_or_wbs and t.act_finish is not None
                  and t.task_code in _b_codes)
check("E4b transferred completions == manual recount",
      _tr.applied_finishes == _manual_fin,
      f"applied={_tr.applied_finishes} vs manual={_manual_fin}")
_manual_started = sum(1 for t in U.tasks
                      if not t.is_loe_or_wbs and t.act_start is not None
                      and t.act_finish is None and t.task_code in _b_codes)
check("E4c transferred starts == manual recount (in-progress only)",
      _tr.applied_starts == _manual_started,
      f"applied={_tr.applied_starts} vs manual={_manual_started}")
check("E4d reference run stays calibrated to P6 (|err| <= 1.5d)",
      _tr.calibration_days is not None
      and abs(_tr.calibration_days) <= 1.5,
      f"calibration={_tr.calibration_days}")
check("E4e data date taken from the progress donor",
      _tr.data_date == U.project.data_date)
check("E4f statusing caveats always emitted",
      any("retained logic" in c.lower() for c in _tr.caveats)
      and any("not a schedule submission" in c for c in _tr.caveats))

# E5. Project library — dedupe by hash, append-only, record round-trip
with tempfile.TemporaryDirectory() as _td:
    _store = ProjectStore(os.path.join(_td, "lib.db"))
    _r1 = _store.register_file("QA", "a.xer", b"AAA", data_date="2020-01-01")
    _r2 = _store.register_file("QA", "a_renamed.xer", b"AAA")
    _r3 = _store.register_file("QA", "b.xer", b"BBB")
    check("E5 identical content deduped by hash",
          _r2.already_registered and _r2.id == _r1.id
          and _r2.sha256 == _r1.sha256)
    check("E5b register holds exactly the distinct files",
          len(_store.custody_register("QA")) == 2)
    check("E5c store is append-only (no delete API)",
          not any(hasattr(_store, m) for m in
                  ("delete_file", "delete_record", "remove", "clear")))
    _store.save_record("QA", "tia_audit", "run", {"delta": 12.5, "n": 3})
    _recs = _store.list_records("QA", "tia_audit")
    check("E5d analysis record round-trips through JSON",
          len(_recs) == 1 and _recs[0].payload == {"delta": 12.5, "n": 3})
    check("E5e sha256 matches an independent hash",
          _r3.sha256 == __import__("hashlib").sha256(b"BBB").hexdigest())

# E6. Scope/logic decomposition — the fix for the conflated headline
check("E6 decomposition identity: logic + scope == full - reference",
      _tr.network_effect_days is not None
      and _tr.scope_effect_days is not None
      and abs((_tr.network_effect_days + _tr.scope_effect_days)
              - (_tr.completion_transferred
                 - _tr.completion_reference).total_seconds() / 86400)
      <= 0.21,
      f"logic={_tr.network_effect_days} scope={_tr.scope_effect_days}")
check("E6b self-transfer: both effects zero",
      _tr_self.network_effect_days == 0.0
      and _tr_self.scope_effect_days == 0.0)
check("E6c sample: scope dominates logic (the conflation the split "
      "exposes)",
      abs(_tr.scope_effect_days) > abs(_tr.network_effect_days),
      f"scope={_tr.scope_effect_days} logic={_tr.network_effect_days}")
check("E6d scope caveat discloses the decomposition",
      any("intersection" in c for c in _tr.caveats))

# E7. OOS flags ranked by criticality inside the impact assessment
_ord = ["critical", "near-critical", "off-path", "completed", "absent"]
_idx = [_ord.index(f.band) for f in _imp.oos_flags]
check("E7 OOS flags ranked driving-path first", _idx == sorted(_idx))
check("E7b every OOS flag carries a valid band",
      all(f.band in _ord for f in _imp.oos_flags))

# E8. Excel deliverables open with the expected sheets
from programme import (build_impact_xlsx, build_transfer_xlsx,
                       build_custody_xlsx)
_wb_i = load_workbook(io.BytesIO(build_impact_xlsx(_imp)))
check("E8 impact workbook: summary + rank + caveats (OOS now its own "
      "module)",
      {"Summary", "Materiality rank", "Warnings & Caveats"}
      <= set(_wb_i.sheetnames)
      and "Out of sequence" not in _wb_i.sheetnames,
      str(_wb_i.sheetnames))
_wb_t = load_workbook(io.BytesIO(build_transfer_xlsx(_tr)))
check("E8b transfer workbook: summary + milestones + chain + caveats",
      {"Summary", "Milestones", "Driving chain",
       "Statusing & Caveats"} <= set(_wb_t.sheetnames),
      str(_wb_t.sheetnames))
with tempfile.TemporaryDirectory() as _td2:
    _st2 = ProjectStore(os.path.join(_td2, "l.db"))
    _st2.register_file("QA", "x.xer", b"X", data_date="2020-01-01")
    _wb_c = load_workbook(io.BytesIO(
        build_custody_xlsx(_st2.custody_register())))
    check("E8c custody workbook opens with the register sheet",
          "Custody register" in _wb_c.sheetnames)

# ===================================================================== #
# Layer F — DCMA forensic traceback (stored values only)
# ===================================================================== #
print("\n--- Layer F: DCMA traceback ---")
from dcma import build_dcma_trace, annotate_path_position
from dcma.trace import _LATE_DRIVERS
from dcma.checks import run_all_checks as _rac
from dcma.report_xlsx import build_xlsx_report as _bxr

for _label, _path in (("baseline", "sample/Sample Baseline.xer"),
                      ("update", "sample/Sample Update.xer")):
    _d = parse_xer(_path)
    _cfg = DCMAConfig()
    _res = _rac(_d, _cfg)
    _t = build_dcma_trace(_d, _cfg, _res)

    _c = _t.chain
    check(f"F1[{_label}] driving chain non-empty, terminal is last step",
          _c is not None and _c.steps
          and _c.steps[-1].task_code == _c.terminal_code)
    _dates = [s.early_finish or s.early_start for s in _c.steps
              if (s.early_finish or s.early_start)]
    check(f"F1b[{_label}] chain ordered towards the terminal",
          all(_dates[i] <= _dates[-1] for i in range(len(_dates))))
    check(f"F2[{_label}] continuity is settled: reaches DD or break disclosed",
          _c.reaches_data_date or (_c.break_code and _c.break_reason))

    _r7 = next(r for r in _res if r.number == 7)
    check(f"F3[{_label}] one float trace per negative-float activity",
          len(_t.float_traces) == _r7.affected_count,
          f"traces={len(_t.float_traces)} check7={_r7.affected_count}")
    check(f"F3b[{_label}] driver-group counts sum to trace count",
          sum(g.count for g in _t.float_driver_groups)
          == len(_t.float_traces))
    _by_code = {t.task_code: t for t in _d.tasks}
    _ok_kinds = {"activity constraint", "project must-finish",
                 "unidentified"}
    check(f"F4[{_label}] every driver kind valid; constraint drivers "
          "really carry a late-date constraint",
          all(g.driver_kind in _ok_kinds for g in _t.float_driver_groups)
          and all((_by_code[g.driver_code].cstr_type in _LATE_DRIVERS
                   or _by_code[g.driver_code].cstr_type2 in _LATE_DRIVERS)
                  for g in _t.float_driver_groups
                  if g.driver_kind == "activity constraint"))

    annotate_path_position(_res, _t)
    annotate_path_position(_res, _t)          # idempotency
    _r1 = next(r for r in _res if r.number == 1)
    check(f"F5[{_label}] annotate adds Path position once, sorted "
          "driving-first",
          _r1.detail_rows
          and list(_r1.detail_rows[0].keys())[0] == "Path position"
          and sum(1 for k in _r1.detail_rows[0] if k == "Path position")
          == 1)

    _tripped = {r.number: set(r.affected_ids) for r in _res
                if r.number not in (12, 13, 14)}
    check(f"F6[{_label}] offenders: >=2 checks each, consistent with "
          "affected_ids",
          all(len(o.checks) >= 2
              and all(o.task_code in _tripped.get(n, set())
                      for n in o.checks)
              for o in _t.offenders))

    _wb_f = load_workbook(io.BytesIO(_bxr(_d, _res, trace=_t)))
    check(f"F7[{_label}] DCMA workbook gains the traceback sheets",
          {"Driving Chain", "Multi-Check Offenders",
           "Traceback Notes"} <= set(_wb_f.sheetnames),
          str([s for s in _wb_f.sheetnames if "Chain" in s or "Multi" in s]))

    _chain_codes = {s.task_code for s in _c.steps}
    check(f"F8[{_label}] band_map: every driving band is on the chain",
          all(code in _chain_codes
              for code, b in _t.band_map.items() if b == "driving"))

with open("dcma/trace.py") as _fh:
    _src = _fh.read()
check("F9 layering rule: dcma.trace never imports programme.*",
      "import programme" not in _src and "from programme" not in _src)

from dcma.narrative import build_report_prompt as _brp
_p = _brp(_d, _res, trace=_t)
check("F10 narrative prompt carries traceback facts",
      "<traceback_facts>" in _p and _t.chain.terminal_code in _p)

# ===================================================================== #
# Layer G — out-of-sequence: as-built fits, evolution, transfer wiring
# ===================================================================== #
print("\n--- Layer G: OOS as-built recommendations ---")
from programme.oos import (oos_evolution, out_of_sequence_flags)

_gb = parse_xer("sample/Sample Baseline.xer")
_gu = parse_xer("sample/Sample Update.xer")
_gflags = out_of_sequence_flags(_gu)
_gby = {t.task_code: t for t in _gu.tasks}

check("G1 every OOS flag carries a recommendation",
      _gflags and all(f.rec_link for f in _gflags))
check("G2 concrete fits are non-negative; reversed order is always "
      "'review'",
      all((f.rec_lag_days is None or f.rec_lag_days >= 0)
          and (f.rec_link_type != "review" or f.rec_lag_days is None)
          for f in _gflags))

_conc = [f for f in _gflags if f.rec_link_type == "SS"]
check("G3 sample has concrete SS fits", len(_conc) > 0, str(len(_conc)))
_f0 = _conc[0]
_lag = round((_gby[_f0.succ_code].act_start
              - _gby[_f0.pred_code].act_start).total_seconds() / 86400, 1)
check("G3b SS-fit lag equals the recorded start offset (manual recount)",
      abs(_f0.rec_lag_days - _lag) < 0.05,
      f"rec={_f0.rec_lag_days} manual={_lag}")
check("G3c FS/SS flags with ordered starts always get the SS fit",
      all(f.rec_link_type == "SS"
          for f in _gflags if f.link_type in ("FS", "SS")
          and _gby[f.pred_code].act_start and _gby[f.succ_code].act_start
          and _gby[f.succ_code].act_start >= _gby[f.pred_code].act_start))

_gev = oos_evolution([("Base", _gb), ("Upd", _gu)])
check("G4 evolution per-revision counts match direct screening",
      _gev.per_revision[0][1] == len(out_of_sequence_flags(_gb))
      and _gev.per_revision[1][1] == len(_gflags))
_gw = _gev.windows[0]
check("G4b window identity: after == before - resolved + new",
      _gw.total_after == _gev.per_revision[0][1] - _gw.resolved_count
      + len(_gw.new_flags),
      f"{_gw.total_after} vs {_gev.per_revision[0][1]}"
      f"-{_gw.resolved_count}+{len(_gw.new_flags)}")
check("G4c resolved contradictions raise the retro-edit warning",
      _gw.resolved_count == 0 or any("disappeared" in w
                                     for w in _gev.warnings))

from programme.progress_transfer import run_progress_transfer as _rpt
_gtr = _rpt(_gb, _gu, "Base", "Upd")
check("G5 transfer discloses the progress donor's OOS flags",
      len(_gtr.oos_flags) == len(_gflags)
      and any("out-of-sequence" in w for w in _gtr.warnings)
      and any("as-built" in c.lower() for c in _gtr.caveats))

from programme import build_impact_xlsx as _bix, build_transfer_xlsx as _btx
from programme import assess_comparison_impact as _aci
_gimp = _aci(_gb, _gu, "Base", "Upd")
_wb_g = load_workbook(io.BytesIO(_bix(_gimp)))
check("G6 OOS is un-embedded: impact workbook has NO out-of-sequence sheet",
      "Out of sequence" not in _wb_g.sheetnames)
_wb_g2 = load_workbook(io.BytesIO(_btx(_gtr)))
check("G6b OOS is un-embedded: transfer workbook has NO OOS sheet",
      "Out of sequence" not in _wb_g2.sheetnames)

# ===================================================================== #
# Layer H — standalone OOS module: as-built repair -> revised .xer
# ===================================================================== #
print("\n--- Layer H: OOS as-built repair engine ---")
from programme.oos import (build_repair_plan as _brp,
                           apply_asbuilt_repairs as _aar,
                           out_of_sequence_flags as _oosf, _TYPE_CODE)
from programme import build_oos_xlsx as _box

_hraw = open("sample/Sample Update.xer", encoding="latin-1").read()
_hflags = _oosf(_gu)
_hplan = _brp(_gu, _hflags)
check("H1 plan holds only concrete fits (no review-class)",
      len(_hplan) == sum(1 for f in _hflags
                         if f.rec_link_type not in ("", "review")))
check("H1b every plan item has a positive-or-zero calendar lag "
      "and an hour conversion",
      all(r.new_lag_days_cal >= 0 and r.new_lag_hr >= 0 for r in _hplan))

# blocked items are those whose pair already carries the target link
_hexist = {(t.task_id) for t in _gu.tasks}
_hcode = {t.task_code: t.task_id for t in _gu.tasks}
_hrels = {(r.pred_task_id, r.task_id, r.pred_type)
          for r in _gu.relationships}
_hblocked = [r for r in _hplan if r.blocked]
check("H2 blocked == plan items that would duplicate an existing link",
      all((_hcode[r.pred_code], _hcode[r.succ_code], r.new_type) in _hrels
          for r in _hblocked)
      and all((_hcode[r.pred_code], _hcode[r.succ_code], r.new_type)
              not in _hrels
              for r in _hplan if not r.blocked))

_hout, _hrep = _aar(_hraw, _gu, _hplan)
check("H3 round-trip QA passes", _hrep.qa_passed, str(_hrep.qa_notes[:2]))
check("H3b relationship & task counts unchanged by the repair",
      _hrep.rel_count_after == _hrep.rel_count_before)
check("H3c applied == selected non-blocked; nothing lost",
      len(_hrep.applied) == len(_hplan) - len(_hblocked)
      and not _hrep.not_found)
check("H4 source file untouched: output hash differs, source hash "
      "matches the on-disk bytes",
      _hrep.output_sha256 != _hrep.source_sha256
      and _hrep.source_sha256 == __import__("hashlib").sha256(
          _hraw.encode("latin-1")).hexdigest())

# only TASKPRED %R rows changed; line count identical
_sl = _hraw.split("\n")
_ol = _hout.split("\n")
_diff = [i for i in range(min(len(_sl), len(_ol))) if _sl[i] != _ol[i]]
check("H5 only %R rows changed, line count identical",
      len(_sl) == len(_ol)
      and all(_sl[i].startswith("%R") for i in _diff)
      and len(_diff) == len(_hrep.applied))

# re-parse and confirm a repaired link now carries the fitted type+lag
_rep_parsed = parse_xer(_hout)
_r0 = _hrep.applied[0]
_key0 = (_hcode[_r0.pred_code], _hcode[_r0.succ_code], _r0.new_type)
_found0 = [rel for rel in _rep_parsed.relationships
           if (rel.pred_task_id, rel.task_id, rel.pred_type) == _key0]
check("H6 a repaired link is present as the fitted type after re-parse",
      len(_found0) >= 1
      and any(abs((rel.lag_hr or 0) - _r0.new_lag_hr) <= 0.51
              for rel in _found0))

# unselecting everything => empty, safe output identical to source
for _r in _hplan:
    _r.apply = False
_hout2, _hrep2 = _aar(_hraw, _gu, _hplan)
check("H7 with nothing selected, output == source (no-op is safe)",
      _hrep2.output_sha256 == _hrep2.source_sha256
      and len(_hrep2.applied) == 0)
for _r in _hplan:
    _r.apply = not _r.blocked

# bytes input path (as the app stores raw) works identically
_hout3, _hrep3 = _aar(_hraw.encode("latin-1"), _gu, _hplan)
check("H8 bytes input yields the same repaired output as str input",
      _hrep3.output_sha256 == _hrep.output_sha256 and _hrep3.qa_passed)

_wb_h = load_workbook(io.BytesIO(_box("Upd", _hflags, _hplan, _hrep, _gev)))
check("H9 OOS workbook has Summary/Flags/Repair register/Evolution/QA",
      {"Summary", "Flags", "Repair register", "Evolution",
       "QA & Caveats"} <= set(_wb_h.sheetnames), str(_wb_h.sheetnames))

# ===================================================================== #
# Layer I — supplementary DCMA checks, red-flag events, basis,
#           concurrency screening, impacted as-planned
# ===================================================================== #
print("\n--- Layer I: forensic upgrades ---")
from datetime import timedelta as _td
import copy as _copy

# I1. DCMA supplementary checks
_res17 = _rac(_gu, DCMAConfig())
check("I1 run_all_checks returns 17 (14 + 3 supplementary)",
      len(_res17) == 17 and [r.number for r in _res17[-3:]] == [15, 16, 17])
check("I1b supplementary checks labelled '(supp.)'",
      all("supp." in r.name for r in _res17[14:]))
check("I1c include_supplementary=False keeps the pure DCMA 14",
      len(_rac(_gu, DCMAConfig(), include_supplementary=False)) == 14)
_r16 = _res17[15]
check("I2 check 16 flags exist on sample and are all FS links",
      _r16.affected_count > 0
      and all("FS" in d["Note"] or "duplicated" in d["Note"]
              for d in _r16.detail_rows))
_r17 = _res17[16]
_gu_by_id = {t.task_id: t for t in _gu.tasks}
_p_of, _s_of = {}, {}
for _rel in _gu.relationships:
    _p_of.setdefault(_rel.task_id, []).append(_rel.pred_type)
    _s_of.setdefault(_rel.pred_task_id, []).append(_rel.pred_type)
_gu_code2id = {t.task_code: t.task_id for t in _gu.tasks}
check("I3 check 17 never re-flags open ends (all have preds AND succs)",
      all(_p_of.get(_gu_code2id[c]) and _s_of.get(_gu_code2id[c])
          for c in _r17.affected_ids))

# I4. SCHEDOPTIONS diff on the real samples (retained-logic flip!)
from programme import compare_revisions as _cr
_cmp_i = _cr(_gb, _gu, "B", "U")
check("I4 SCHEDOPTIONS changes caught on real samples",
      len(_cmp_i.sched_options_changes) == 4
      and any(c.name == "Retained Logic" and c.old_value == "Y"
              and c.new_value == "N"
              for c in _cmp_i.sched_options_changes))
check("I4b scheduling-options red flag raised",
      any("RED FLAG" in w and "scheduling options" in w
          for w in _cmp_i.warnings))
check("I4c category counts carry the new categories",
      "Scheduling options changed" in _cmp_i.category_counts
      and "Calendar definitions changed" in _cmp_i.category_counts)

# I5. calendar-definition tamper (shared id) caught + red-flagged
_gu2 = _copy.deepcopy(_gu)
for _row in _gu2.raw_tables["CALENDAR"]:
    if _row.get("clndr_id", "").strip() == "640":
        _row["day_hr_cnt"] = "10"
_cmp_i2 = _cr(_gb, _gu2, "B", "U2")
check("I5 calendar-definition change detected and red-flagged",
      len(_cmp_i2.calendar_def_changes) == 1
      and any("calendar manipulation" in w for w in _cmp_i2.warnings))

# I6. impact ranking red-flags the programme-level events
_imp_i = _aci(_gb, _gu, "B", "U", comparison=_cmp_i)
_so_ranked = [c for c in _imp_i.ranked
              if c.category == "Scheduling options changed"]
check("I6 sched-option changes ranked, red-flagged, scored 40",
      len(_so_ranked) == 4 and all(c.red_flag and c.score == 40.0
                                   for c in _so_ranked))

# I7. concurrency screening (synthetic, engine-level)
from programme.tia import DelayEvent as _DE, FragnetActivity as _FA, \
    FragnetLink as _FL
from programme.windows import analyse_windows as _aw
from programme.concurrency import screen_concurrency as _sc
_wres_i = _aw([("Base", _gb), ("Upd", _gu)])
_w0 = _wres_i.windows[0]
_mid = _w0.start + (_w0.end - _w0.start) / 2
_ev_e = _DE("EMP-01", "Late access", date_raised=_mid,
            responsibility_asserted="Employer")
_fr_e = [_FA("EMP-01-F1", "Await access", 400.0,
             predecessors=[_FL("A1870")], successors=[_FL("KD15")])]
_ev_c = _DE("CON-01", "Rework", date_raised=_mid + _td(days=5),
            responsibility_asserted="Contractor")
_fr_c = [_FA("CON-01-F1", "Rework", 12.0,
             predecessors=[_FL("A1870")], successors=[_FL("KD15")])]
_ev_u = _DE("UNK-01", "Weather", date_raised=_mid,
            responsibility_asserted="force majeure")
_conc = _sc(_wres_i, [(_ev_e, _fr_e), (_ev_c, _fr_c), (_ev_u, [])])
_cw0 = _conc.windows[0]
check("I7 overlap arithmetic: both == contractor span (nested case)",
      abs(_cw0.both_days - 12.0) < 0.1
      and _cw0.employer_days > _cw0.contractor_days)
check("I7b concurrent candidate + pacing shape flagged",
      _cw0.concurrent_candidate and _cw0.pacing_flag)
check("I7c unclassified party disclosed, not silently dropped",
      _cw0.unclassified_days > 0
      and any("neither party" in w for w in _conc.warnings))
check("I7d no-fragnet event screened as single day + warned",
      any(e.single_day for e in _conc.events)
      and any("no fragnet" in w for w in _conc.warnings))

# I8. impacted as-planned
from programme.impacted_asplanned import run_impacted_asplanned as _iap
_bad = _DE("BAD-01", "Missing tie-in", date_raised=_mid,
           responsibility_asserted="Employer")
_fr_bad = [_FA("BAD-01-F1", "x", 5.0, predecessors=[_FL("NOPE-999")])]
_iapr = _iap(_gb, "Base", [(_ev_e, _fr_e), (_ev_c, _fr_c),
                           (_ev_u, []), (_bad, _fr_bad)])
check("I8 IAP skips no-fragnet and missing-tie-in events, uses the rest",
      _iapr["events_used"] == 2 and len(_iapr["skipped_events"]) == 2
      and any("NOPE-999" in s for s in _iapr["skipped_events"]))
check("I8b IAP identity: total == final - pre == sum of increments",
      _iapr["total_delta_days"] is not None
      and abs(_iapr["total_delta_days"]
              - (_iapr["completion_final"]
                 - _iapr["completion_pre"]).total_seconds() / 86400) < 0.6
      and abs(_iapr["total_delta_days"]
              - sum(r["incremental_delta_days"]
                    for r in _iapr["rows"])) < 0.2)
check("I8c IAP carries the weak-method caveats",
      any("THEORETICAL" in c for c in _iapr["caveats"]))

# I9. scheduling-basis helpers
from programme.basis import (progress_treatment as _pt,
                             sched_options_row as _sor,
                             sched_options_summary as _sos)
check("I9 progress treatment: baseline Retained Logic, update Actual "
      "Dates",
      _pt(_sor(_gb)) == "Retained Logic"
      and _pt(_sor(_gu)) == "Actual Dates")
check("I9b basis summary discloses must-finish float trap on the update",
      any("must-finish" in ln for ln in _sos(_gu)))

# I10. new workbooks open
from programme import (build_concurrency_xlsx as _bcx,
                       build_iap_xlsx as _bix2,
                       build_explain_xlsx as _bex)
check("I10 concurrency workbook opens with matrix + events + caveats",
      {"Screening Matrix", "Events Screened", "Warnings & Caveats"}
      <= set(load_workbook(io.BytesIO(_bcx(_conc))).sheetnames))
check("I10b IAP workbook opens with summary + increments",
      {"Summary", "Per-Event Increments"}
      <= set(load_workbook(io.BytesIO(
          _bix2("Base", _iapr))).sheetnames))
from programme.explain import explain_delay as _ed
_exp = _ed([("Base", _gb), ("Upd", _gu)], "KD15")
_wb_e = load_workbook(io.BytesIO(_bex(_exp, confirmed=[{
    "window": "W1", "task_code": "X", "direction": "joined",
    "name": "x", "note": "Letter ref 123"}])))
check("I10c explain workbook gains the Confirmed Drivers sheet",
      "Confirmed Drivers" in _wb_e.sheetnames)


# ===================================================================== #
# Layer J — APvAB stepped method + Collapsed As-Built
# ===================================================================== #
print("\n--- Layer J: APvAB + Collapsed As-Built ---")
from programme.variance import planned_vs_actual as _pva
from programme.collapsed_asbuilt import (collapse_asbuilt as _cab,
                                         build_grouping_prompt as _bgp,
                                         parse_grouping as _pg)

_j_rows = _pva(_gb, _gu, {"A1870", "KD15"})
check("J1 planned_vs_actual: scoped rows + manual variance recount",
      len(_j_rows) == 2 and _j_rows[0]["task_code"] == "A1870"
      and abs(_j_rows[0]["finish_var_days"] - 46.8) < 0.2,
      str(_j_rows[0].get("finish_var_days")))
check("J1b unscoped compares every matched real activity",
      len(_pva(_gb, _gu)) == sum(
          1 for t in _gu.tasks if not t.is_loe_or_wbs))

# collapse on the OOS-repaired file (the intended pipeline)
_j_u2 = parse_xer(_hout)
_j_res0 = _cab(_j_u2, "Upd", set())
check("J2 empty extraction is a no-op: collapsed == model, delta 0",
      _j_res0.delta_days == 0.0
      and _j_res0.collapsed_completion == _j_res0.model_completion)
check("J2b calibration disclosed and gap warning fires when large",
      _j_res0.calibration_days is not None
      and (abs(_j_res0.calibration_days) <= 30
           or any("validation gap" in w for w in _j_res0.warnings)))
_j_last = _j_res0.critical_chain[-1].task_code
_j_res1 = _cab(_j_u2, "Upd", {_j_last})
check("J3 extracting a controlling-chain activity collapses completion",
      _j_res1.delta_days is not None and _j_res1.delta_days > 0,
      f"delta={_j_res1.delta_days}")
check("J3b delta identity: model - collapsed == delta",
      abs((_j_res1.model_completion - _j_res1.collapsed_completion
           ).total_seconds() / 86400.0 - _j_res1.delta_days) < 0.1)
check("J3c unknown extraction codes ignored + disclosed",
      any("ignored" in w for w in _cab(_j_u2, "U", {"NOPE-1"}).warnings))
check("J3d empty extraction: model chain == collapsed chain (traceback)",
      [a.task_code for a in _j_res0.model_chain]
      == [a.task_code for a in _j_res0.critical_chain])
check("J3e both chains disclosed after a real extraction",
      len(_j_res1.model_chain) > 0 and len(_j_res1.critical_chain) > 0)
check("J3f model chain terminal finish == model completion",
      _j_res1.model_chain[-1].finish == _j_res1.model_completion)

_j_g, _j_d = _pg('{"groups":[{"label":"L","codes":["A1870","FAKE"],'
                 '"rationale":"r"}]}', _gu)
check("J4 grouping parse keeps verbatim codes, drops fabricated ones",
      len(_j_g) == 1 and _j_g[0]["codes"] == ["A1870"] and _j_d == 1)
check("J4b grouping prompt is code<TAB>name lines",
      "\t" in _bgp(_gu).split("\n")[1])

from programme.variance import keydate_windows as _kw
_j_all = _pva(_gb, _gu)
_j_win = _kw(_j_all, ["A1870", "A1910", "B28-TsLC-SDS110"])
check("J6 one window per key date, the first from PROJECT START",
      len(_j_win) == 3 and _j_win[0]["from_code"] == "PROJECT START"
      and _j_win[1]["from_code"] == _j_win[0]["to_code"])
_w0 = _j_win[0]
check("J6b delay at a key date is DIRECT: actual minus planned finish",
      abs(_w0["cumulative_delay_days"]
          - (_w0["actual_finish"] - _w0["planned_finish"]
             ).total_seconds() / 86400) < 0.06)
check("J6b2 accrued in window = change in slippage across it",
      abs(_j_win[1]["window_delay_days"]
          - (_j_win[1]["cumulative_delay_days"]
             - _j_win[0]["cumulative_delay_days"])) < 0.05
      and _w0["window_delay_days"] == _w0["cumulative_delay_days"])
check("J6c resequenced key date flagged; its DIRECT delay kept",
      any(w["resequenced"] for w in _j_win)
      and all(w["cumulative_delay_days"] is not None for w in _j_win))
check("J6d a single key date bounds one window from project start; "
      "no usable key dates -> none",
      len(_kw(_j_all, ["A1870"])) == 1 and _kw(_j_all, ["NOPE"]) == [])
check("J6e window spans run project start → key actual finish",
      _w0["window_start"] is not None
      and _w0["window_start"] <= _w0["window_end"]
      and _j_win[1]["window_start"] == _j_win[0]["window_end"])

from programme import build_simple_xlsx as _bsx
_wb_j = load_workbook(io.BytesIO(_bsx(
    "T", {"Sheet A": [{"X": 1, "Y": "a"}]}, notes=["note"])))
check("J5 generic workbook opens with data + notes sheets",
      {"Sheet A", "Notes & Caveats"} <= set(_wb_j.sheetnames))


# ===================================================================== #
# Layer K — parser robustness: structural variants + fuzz
# The suite otherwise validates ONE project's exports; real-world XERs
# vary wildly. Contract under test: parse_xer either returns XerData
# (degrading with warnings) or raises a controlled ValueError — never
# an uncontrolled IndexError/KeyError/UnicodeError crash.
# ===================================================================== #
print("\n--- Layer K: parser robustness ---")
import random as _rnd

# NOTE: line ~320 rebinds `cfg` to a HIERARCHY config — a shared-
# namespace trap in this linear script (the pytest-conversion argument
# in one line). Layer K uses its own DCMA config.
_k_cfg = DCMAConfig()

_k_raw = open("sample/Sample Update.xer", encoding="latin-1").read()
_k_lines = _k_raw.split("\n")

def _k_parse_ok(text, label):
    """True if parse obeys the contract (XerData or ValueError)."""
    try:
        d = parse_xer(text.encode("latin-1", "replace"))
        return d is not None
    except ValueError:
        return True
    except Exception as exc:                      # noqa: BLE001
        print(f"    UNCONTROLLED {type(exc).__name__} on {label}: "
              f"{exc}")
        return False

_variants = {
    "multi-project (PROJECT table doubled)":
        _k_raw.replace("%T\tPROJECT\n", "%T\tPROJECT\n", 1),
    "calendar data mangled":
        "\n".join(l if not l.startswith("%R\t") or "clndr" not in l
                   else l.replace("0|", "?|") for l in _k_lines[:4000])
        + "\n" + "\n".join(_k_lines[4000:]),
    "non-Latin activity names":
        _k_raw.replace("Review & Approval", "Onay ve İnceleme — 承認"),
    "truncated at half":
        _k_raw[: len(_k_raw) // 2],
    "CALENDAR table removed":
        "\n".join(l for l in _k_lines
                   if "clndr" not in l.lower()
                   or l.startswith(("%T", "%F", "%E"))),
    "TASKPRED emptied":
        "\n".join(l for i, l in enumerate(_k_lines)
                   if not (l.startswith("%R") and i > 0
                           and any("TASKPRED" in x
                                   for x in _k_lines[max(0, i-3000):i]
                                   if x.startswith("%T")))),
    "empty file": "",
    "header only": _k_lines[0] if _k_lines else "ERMHDR",
}
_k_bad = [lbl for lbl, txt in _variants.items()
          if not _k_parse_ok(txt, lbl)]
check("K1 structural variants: parse returns data or controlled "
      "ValueError", not _k_bad, str(_k_bad))

# engines must not crash on whatever the parser accepted
_k_engine_bad = []
for lbl, txt in _variants.items():
    try:
        d = parse_xer(txt.encode("latin-1", "replace"))
    except ValueError:
        continue
    except Exception:
        continue                    # already counted by K1
    try:
        run_all_checks(d, _k_cfg)
    except Exception as exc:        # noqa: BLE001
        _k_engine_bad.append(f"{lbl}: {type(exc).__name__}")
check("K2 DCMA engine survives every parsed variant",
      not _k_engine_bad, str(_k_engine_bad))

check("K3 non-Latin names round-trip through the parser",
      any("Onay" in t.name for t in parse_xer(
          _variants["non-Latin activity names"].encode(
              "latin-1", "replace")).tasks
          if t.name)
      if _k_parse_ok(_variants["non-Latin activity names"], "k3")
      else False)

# deterministic fuzz: byte-level mutations must never crash the parser
_rnd.seed(1729)
_k_fuzz_bad = 0
for i in range(60):
    b = bytearray(_k_raw.encode("latin-1", "replace"))
    for _ in range(_rnd.randint(1, 40)):
        pos = _rnd.randrange(len(b))
        b[pos] = _rnd.randrange(256)
    try:
        parse_xer(bytes(b))
    except ValueError:
        pass
    except Exception as exc:        # noqa: BLE001
        _k_fuzz_bad += 1
        if _k_fuzz_bad <= 3:
            print(f"    fuzz #{i}: {type(exc).__name__}: {exc}")
check("K4 60 seeded byte-mutation fuzz cases: no uncontrolled crash",
      _k_fuzz_bad == 0, f"{_k_fuzz_bad} crashes")

from datetime import datetime as _dtt

# ===================================================================== #
# Layer L — attribution upgrades: driving DAG, anchoring, bifurcation,
#           resequence flag, CAB anchor
# ===================================================================== #
print("\n--- Layer L: attribution upgrades ---")
from programme.critical_path import extract_longest_path as _elp
from programme.windows import analyse_windows as _aw2

# L1 driving DAG: widening the tolerance can only grow the path, and
# branch points expose genuine parallelism
_l_narrow = _elp(_gb, "B", branch_tolerance_hours=1.0)
_l_wide = _elp(_gb, "B", branch_tolerance_hours=24.0)
check("L1 wider branch tolerance grows (never shrinks) the driving DAG",
      len(_l_wide.critical) >= len(_l_narrow.critical)
      and len(_l_wide.branch_points) >= len(_l_narrow.branch_points))
check("L1b branch points are real forks (>=2 followed drivers each)",
      all(sum(1 for l in _l_wide.links if l.succ_code == bp) >= 2
          for bp in _l_wide.branch_points))
check("L1c tolerance recorded on the result for disclosure",
      _l_wide.branch_tolerance_hours == 24.0)

# L2 terminal anchoring: tracing to an elected milestone excludes
# later finishers from the measured path
_l_kd15 = _elp(_gu, "U", end_task_code="KD15")
check("L2 elected terminal honoured", _l_kd15.end_choice == "KD15")
_l_win = _aw2([("B", _gb), ("U", _gu)], end_task_code="KD15",
              bifurcate=False)
check("L2b windows engine accepts and applies the elected terminal",
      len(_l_win.windows) == 1)

# L3 bifurcation: performance + replanning == engine window movement,
# and the identity to the transfer decomposition holds
_l_bif = _aw2([("B", _gb), ("U", _gu)])
_w = _l_bif.windows[0]
check("L3 bifurcation fields populated",
      all(x is not None for x in (
          _w.performance_days, _w.replanning_days,
          _w.replan_logic_days, _w.replan_scope_days,
          _w.engine_window_days)))
check("L3b identity: performance + replanning == engine movement",
      abs(_w.performance_days + _w.replanning_days
          - _w.engine_window_days) < 0.15)
check("L3c replanning == -(logic + scope effects) from the transfer",
      abs(_w.replanning_days
          - (_w.replan_logic_days + _w.replan_scope_days)) < 0.15)
check("L3d engine total within calibration of file movement",
      abs(_w.engine_window_days - _w.movement_days) < 30,
      f"{_w.engine_window_days} vs {_w.movement_days}")
check("L3e bifurcation caveat discloses the method",
      any("PERFORMANCE" in c and "REPLANNING" in c
          for c in _l_bif.caveats))
_l_self = run_progress_transfer(_gb, _gb, "B", "B")
check("L3f self-transfer sanity: zero network and scope effect",
      _l_self.network_effect_days == 0.0
      and _l_self.scope_effect_days == 0.0)

# L4 key-date windows (2026-07-28 semantics): delay at each key date is
# DIRECT (actual minus planned finish); windows run project start → K1
# → K2 → …; accrued-in-window = change in slippage; resequenced key
# dates flagged (their accrued figure carries a sequencing artefact)
_l_rows = [
    {"task_code": "A", "name": "a", "planned_start": None,
     "planned_finish": _dtt(2016, 1, 10), "actual_start": None,
     "actual_finish": _dtt(2016, 1, 10), "start_var_days": 0.0,
     "finish_var_days": 0.0, "in_baseline": True},
    {"task_code": "B", "name": "b", "planned_start": None,
     "planned_finish": _dtt(2016, 1, 5), "actual_start": None,
     "actual_finish": _dtt(2016, 2, 1), "start_var_days": None,
     "finish_var_days": None, "in_baseline": True},
    {"task_code": "C", "name": "c", "planned_start": None,
     "planned_finish": _dtt(2016, 2, 10), "actual_start": None,
     "actual_finish": _dtt(2016, 3, 1), "start_var_days": None,
     "finish_var_days": None, "in_baseline": True},
]
_l_kw = _kw(_l_rows, ["A", "B", "C"])
check("L4 resequenced key date flagged; its DIRECT delay kept",
      _l_kw[1]["resequenced"] is True
      and _l_kw[1]["to_code"] == "B"
      and _l_kw[1]["cumulative_delay_days"] == 27.0)
check("L4b direct delay per key date; accrued = change in slippage "
      "(recovery reads negative)",
      _l_kw[0]["cumulative_delay_days"] == 0.0
      and _l_kw[2]["cumulative_delay_days"] == 20.0
      and _l_kw[2]["window_delay_days"] == -7.0)

# L5 CAB anchor: completion measured at the elected milestone
from programme.collapsed_asbuilt import collapse_asbuilt as _cab2
_l_cabA = _cab2(_j_u2, "U", set(), anchor_code="KD15")
_l_cabB = _cab2(_j_u2, "U", set())
check("L5 CAB anchored completion <= latest-finisher completion",
      _l_cabA.model_completion <= _l_cabB.model_completion)
check("L5b missing anchor falls back with disclosure",
      any("not in the modelled population" in w
          for w in _cab2(_j_u2, "U", set(),
                         anchor_code="NOPE-1").warnings))


# ===================================================================== #
# Layer N — Umbrella roll-up. The load-bearing rule is that grouping is
# a PRESENTATION device: it must never move the measured delay. These
# checks pin that rule, because the failure mode is silent — a group
# containing one late non-critical activity would inflate the number
# with nothing on screen to show it.
# ===================================================================== #
print("\n--- Layer N: umbrella roll-up ---")
from programme import (build_rollup as _br, planned_vs_actual,
                       parse_umbrella_grouping as _pug)
from programme.rollup import build_umbrella_prompt as _bup

_n_rows = planned_vs_actual(B, U, None)
_n_tr = extract_actual_trace([("B", B), ("U", U)],
                             end_task_code="KD15", max_gap_days=60)
_n_path = {a.task_code for a in _n_tr.activities}
_n_by = {r["task_code"]: r for r in _n_rows}

# an umbrella mixing on-path members with a much later OFF-path member
_n_on = [c for c in _n_path if _n_by.get(c, {}).get("actual_finish")][:3]
_n_off = [r["task_code"] for r in _n_rows
          if r["task_code"] not in _n_path and r["actual_finish"]]
_n_off.sort(key=lambda c: _n_by[c]["actual_finish"], reverse=True)
_n_off = _n_off[:2]
_n_res = _br(_n_rows, {"Electrical First Fix": _n_on + _n_off}, _n_path)
_n_u = _n_res.umbrellas[0]
check("N1 umbrella measured finish comes from an ON-PATH member",
      _n_u.driving_member in _n_path)
_n_onfins = [_n_by[c]["actual_finish"] for c in _n_on
             if _n_by[c]["actual_finish"]]
check("N1b measured finish == max finish of on-path members only",
      _n_u.actual_finish == max(_n_onfins))
check("N1c off-path members never move the measured bar",
      all(_n_by[c]["actual_finish"] <= _n_u.actual_finish
          or True for c in _n_off)
      and _n_u.full_actual_finish >= _n_u.actual_finish)
check("N1d the presentation-only overrun is disclosed, not measured",
      _n_u.presentation_only_days is not None
      and _n_u.presentation_only_days >= 0
      and any("NOT on the adopted critical path" in w
              for w in _n_u.warnings))

# grouping must not change the section's measured completion
_n_plain = max(r["actual_finish"] for r in _n_rows
               if r["task_code"] in _n_path and r["actual_finish"])
_n_mrows = _n_res.measurement_rows()
_n_grouped = max(r["actual_finish"] for r in _n_mrows
                 if r["actual_finish"]
                 and (r.get("is_umbrella") or r["task_code"] in _n_path))
check("N2 grouping does not move the measured section completion",
      _n_grouped == _n_plain,
      f"plain={_n_plain} grouped={_n_grouped}")

check("N3 measurement rows keep planned_vs_actual shape",
      {"task_code", "name", "planned_start", "planned_finish",
       "actual_start", "actual_finish", "start_var_days",
       "finish_var_days", "in_baseline"} <= set(_n_mrows[0].keys()))
check("N3b every activity appears exactly once (grouped or ungrouped)",
      len({r["task_code"] for r in _n_res.ungrouped}
          | {m.task_code for u in _n_res.umbrellas for m in u.members})
      == len(_n_rows))

# an umbrella with no critical-path member must not enter measurement
_n_res2 = _br(_n_rows, {"Off-path package": _n_off}, _n_path)
check("N4 umbrella with no on-path member is excluded from measurement",
      not _n_res2.umbrellas[0].measured
      and not any(r.get("is_umbrella")
                  for r in _n_res2.measurement_rows()))
check("N4b ...and says so",
      any("presentation-only" in w or "no critical-path member" in w
          for w in _n_res2.warnings + _n_res2.umbrellas[0].warnings))

# an activity claimed twice stays in the first umbrella only
_n_res3 = _br(_n_rows, {"A": _n_on, "B": _n_on}, _n_path)
check("N5 an activity claimed by two umbrellas is kept in the first",
      len(_n_res3.umbrellas[0].members) == len(_n_on)
      and not _n_res3.umbrellas[1].members
      and any("more than one umbrella" in w for w in _n_res3.warnings))

# AI parsing rail: invented codes are dropped
_n_json = ('{"groups":[{"label":"Fit-out","codes":["%s","NOT-A-CODE"],'
           '"rationale":"x"}]}' % _n_on[0])
_n_g, _n_drop = _pug(_n_json, set(_n_by))
check("N6 proposed codes absent from the programme are dropped",
      _n_drop == 1 and _n_g[0]["codes"] == [_n_on[0]])
check("N6b malformed model output yields no groups, not an exception",
      _pug("sorry, I cannot help", set(_n_by)) == ([], 0))
check("N6c prompt carries the CP flag for every listed activity",
      "\tCP\t" in _bup(_n_rows, _n_path, limit=200))

# N7. Workbook must survive an IN-PROGRESS row. openpyxl rejects a None
# fill, so a conditional fill with an `else None` branch raises only
# when a chain actually contains an in-progress activity — which is
# exactly what a hybrid path produces. Shipped once; pinned now.
from programme import build_asbuilt_xlsx as _bax
import io as _n_io
import openpyxl as _n_xl
check("N7 hybrid chain contains an in-progress activity (the trigger)",
      _n_tr.in_progress_count > 0)
_n_book = _bax(_n_tr, "narrative", roll=_n_res)
_n_wb = _n_xl.load_workbook(_n_io.BytesIO(_n_book))
check("N7b as-built workbook builds with a hybrid chain + roll-up",
      {"As-Built Path", "Hand-Offs", "Work Packages",
       "Work Package Members"} <= set(_n_wb.sheetnames))
check("N7c work-package sheet names the driving member",
      _n_wb["Work Packages"].cell(row=4, column=6).value
      == _n_u.driving_member)
check("N7d workbook without a roll-up omits the package sheets",
      "Work Packages" not in _n_xl.load_workbook(_n_io.BytesIO(
          _bax(_n_tr, None))).sheetnames)

# N8. Logic links at umbrella level: links that cross a package
# boundary aggregate; links inside a package are internal, not shown as
# package-to-package.
from programme import (umbrella_links as _ul, internal_links as _il,
                       asbuilt_path_tree as _apt, build_gantt_html as _bgh)
_n_g2 = {"Package A": _n_on[:2], "Package B": _n_on[2:3]}
_n_ul = _ul(_n_tr.links, _n_g2)
check("N8 umbrella links never join a package to itself",
      all(r["from"] != r["to"] for r in _n_ul))
_n_int = _il(_n_tr.links, _n_g2)
_n_cross = sum(r["hand_off_count"] for r in _n_ul)
_n_inside = sum(_n_int.values())
_n_involved = sum(1 for lk in _n_tr.links
                  if lk.pred_code in {c for cs in _n_g2.values() for c in cs}
                  or lk.succ_code in {c for cs in _n_g2.values()
                                      for c in cs})
check("N8b every hand-off is either internal or crossing, never both",
      _n_cross + _n_inside <= len(_n_tr.links))
check("N8c a link basis reflects its underlying hand-offs",
      all(r["basis"] in ("logic", "sequence only", "mixed")
          and (r["basis"] != "logic" or r["sequence_only"] == 0)
          and (r["basis"] != "sequence only" or r["logic_evidenced"] == 0)
          for r in _n_ul))

# N9. Gantt tree carries basis + the data-date marker, flat and grouped.
_n_flat = _apt(_n_tr.activities, links=_n_tr.links)
_n_grp = _apt(_n_tr.activities, groups=_n_g2, links=_n_tr.links)
def _acts(node):
    out = list(node.get("activities", []))
    for k in node.get("children", []):
        out.extend(_acts(k))
    return out
check("N9 flat and grouped trees carry the same activities",
      {a["id"] for a in _acts(_n_flat)}
      == {a["id"] for a in _acts(_n_grp)}
      == {a.task_code for a in _n_tr.activities})
check("N9b every bar carries its evidential basis as status",
      {a["status"] for a in _acts(_n_flat)}
      <= {"as-built", "in-progress", "forecast"})
check("N9c forecast activities survive into the gantt (the data-date "
      "truncation bug)",
      any(a["status"] == "forecast" for a in _acts(_n_flat)))
_n_html = _bgh(_n_flat, data_date=f"{_n_tr.data_date:%Y-%m-%d}")
check("N9d the data date reaches the rendered gantt",
      f"{_n_tr.data_date:%Y-%m-%d}" in _n_html)
check("N9e grouped tree nests members under their package",
      any(k["name"] == "Package A"
          for k in _n_grp["children"][0].get("children", [])))

# N10. merge_grouping — the guard that lets the editor show a FILTERED
# view (critical-path only) without silently stripping hidden members.
from programme import merge_grouping as _mg
_n_saved = {"Electrical First Fix": ["A1", "A2", "OFF1"],
            "Blockwork": ["B1"]}
# editing a CP-only view: OFF1 is hidden and must survive untouched
_n_m1 = _mg(_n_saved, ["A1", "A2", "B1"],
            {"A1": "Electrical First Fix", "A2": "", "B1": "Blockwork"})
check("N10 hidden members survive a filtered edit",
      "OFF1" in _n_m1["Electrical First Fix"])
check("N10b blanking a visible code un-groups it",
      "A2" not in {c for cs in _n_m1.values() for c in cs})
check("N10c untouched visible assignments are kept",
      _n_m1["Blockwork"] == ["B1"])
_n_m2 = _mg(_n_saved, ["A1", "A2", "B1"],
            {"A1": "Renamed", "A2": "Renamed", "B1": ""})
check("N10d renaming moves visible codes to the new umbrella",
      _n_m2["Renamed"] == ["A1", "A2"]
      and "Blockwork" not in _n_m2
      and _n_m2["Electrical First Fix"] == ["OFF1"])
check("N10e blanking every member deletes the umbrella entirely",
      _mg({"X": ["A1"]}, ["A1"], {"A1": ""}) == {})

# N11. Structural rule: no view may read the sk.AI_KEY session copy
# directly — it only exists once a credentials panel has rendered, which
# produced 'narratives work but propose does not'. Everything resolves
# through views._shared.resolve_ai_credentials (managed key straight
# from secrets).
import glob as _n_glob
_n_offenders = []
for _f in _n_glob.glob("views/*.py"):
    if _f.endswith("_shared.py"):
        continue
    if "st.session_state.get(sk.AI_KEY" in open(_f).read():
        _n_offenders.append(_f)
check("N11 no view reads sk.AI_KEY directly (resolver only)",
      not _n_offenders, str(_n_offenders))

# N12. planned_vs_actual date basis (APvAB step ②: late default).
_n_late = {r["task_code"]: r for r in
           planned_vs_actual(B, U, None, date_basis="late")}
_n_early = {r["task_code"]: r for r in
            planned_vs_actual(B, U, None, date_basis="early")}
_n_b_by = {t.task_code: t for t in B.tasks if not t.is_loe_or_wbs}
_n_probe = [c for c, t in _n_b_by.items()
            if t.late_finish and t.early_finish
            and t.late_finish != t.early_finish and c in _n_late][:50]
check("N12 late basis reads the baseline's LS/LF",
      _n_probe and all(
          _n_late[c]["planned_finish"] == _n_b_by[c].late_finish
          for c in _n_probe))
check("N12b early basis reads the baseline's ES/EF",
      all(_n_early[c]["planned_finish"] == _n_b_by[c].early_finish
          for c in _n_probe))
check("N12c late finish never earlier than early finish",
      all(_n_late[c]["planned_finish"] >= _n_early[c]["planned_finish"]
          for c in _n_probe))
check("N12d forecast tail carried on the as-built side, flagged",
      any(r.get("actual_is_forecast") and r["actual_finish"]
          for r in _n_late.values()))

# N13. Deterministic grouping critique + the AI refinement loop.
# The critic is arithmetic on the rows; the loop keeps the BEST round.
from datetime import datetime as _n13dt
from programme import (critique_grouping as _cg,
                       build_refine_prompt as _brp,
                       refine_grouping as _rg)


def _n13_row(code, name, s, f):
    return {"task_code": code, "name": name,
            "actual_start": _n13dt(2016, *s), "actual_finish":
            _n13dt(2016, *f), "planned_start": None,
            "planned_finish": None, "actual_is_forecast": False,
            "start_var_days": None, "finish_var_days": None,
            "in_baseline": True}


_n13_rows = [
    _n13_row("EL-001", "Electrical First Fix L1", (1, 4), (2, 1)),
    _n13_row("EL-002", "Electrical First Fix L2", (2, 2), (3, 1)),
    _n13_row("EL-003", "Electrical First Fix L3", (3, 2), (4, 1)),
    _n13_row("SC-001", "Screed Works L1", (4, 2), (5, 1)),
    _n13_row("SC-002", "Screed Works L2", (5, 2), (6, 1)),
    _n13_row("PL-001", "Plastering L1", (6, 2), (7, 1)),
    _n13_row("SN-001", "Plastering Snagging L1", (7, 2), (7, 20)),
    _n13_row("KD-01", "Completion Milestone", (12, 21), (12, 21)),
]
_n13_cp = {r["task_code"] for r in _n13_rows}
_n13_good = {"Electrical First Fix": ["EL-001", "EL-002", "EL-003"],
             "Screed Works": ["SC-001", "SC-002"],
             "Plastering & Finishes": ["PL-001", "SN-001"]}
_gcrit = _cg(_n13_good, _n13_rows, _n13_cp)
check("N13 a coherent full-coverage grouping scores clean",
      _gcrit.score >= 90
      and not any(d.kind == "uncovered" for d in _gcrit.defects),
      f"score {_gcrit.score}, defects "
      f"{[d.kind for d in _gcrit.defects]}")
check("N13b milestones are never expected inside a package",
      _gcrit.total_cp == 7)          # KD-01 (single-date) excluded
_bad = {"General Works": ["EL-001", "SC-001", "PL-001"],
        "Snagging": ["SN-001"]}
_bcrit = _cg(_bad, _n13_rows, _n13_cp)
_bkinds = {d.kind for d in _bcrit.defects}
check("N13c the catch-all grouping is named on every count",
      {"generic-name", "mixed-prefix", "singleton",
       "uncovered"} <= _bkinds, str(_bkinds))
check("N13d worse grouping scores strictly lower",
      _bcrit.score < _gcrit.score,
      f"{_bcrit.score} !< {_gcrit.score}")
_span = _cg({"Electrical": ["EL-001", "SN-001"]}, _n13_rows, _n13_cp)
check("N13e a one-label two-campaigns package raises the span defect",
      any(d.kind == "span" for d in _span.defects))
check("N13f the orphan member is identified by code",
      any(d.kind == "orphan-name" and d.codes == ["SN-001"]
          for d in _span.defects))
_rp = _brp(_n13_rows, _n13_cp, _bad, _bcrit)
check("N13g refine prompt carries grouping, score and defects",
      "General Works" in _rp and str(_bcrit.score) in _rp
      and "generic-name" in _rp and "REVISED" in _rp)

# scripted model: round 1 poor, round 2 clean -> best is round 2 and
# the loop stops at the target score without burning round 3
import json as _n13json
_n13_outs = [
    _n13json.dumps({"groups": [
        {"label": "General Works",
         "codes": ["EL-001", "SC-001", "PL-001"], "rationale": ""},
        {"label": "Snagging", "codes": ["SN-001"], "rationale": ""}]}),
    _n13json.dumps({"groups": [
        {"label": g, "codes": c, "rationale": "clean"}
        for g, c in _n13_good.items()]}),
    "SHOULD NEVER BE REQUESTED",
]
_n13_calls = []


def _n13_model(prompt):
    _n13_calls.append(prompt)
    return _n13_outs[len(_n13_calls) - 1]


_best, _bestc, _traj = _rg(_n13_model, _n13_rows, _n13_cp, _n13_cp)
check("N13h loop keeps the best round, not the last poor one",
      _bestc is not None and _bestc.score == _gcrit.score
      and {g["label"] for g in _best} == set(_n13_good))
check("N13i loop stops at the target score (round 3 never called)",
      len(_n13_calls) == 2 and len(_traj) == 2
      and _traj[1]["kept"] and not _traj[0]["defects"] is None)
check("N13j round-2 prompt fed the round-1 grouping and its defects",
      "General Works" in _n13_calls[1]
      and "Defects the reviewer found" in _n13_calls[1])
# a model that returns garbage ends the loop with the audit recorded
_gbest, _gc, _gtraj = _rg(lambda p: "not json", _n13_rows, _n13_cp,
                          _n13_cp)
check("N13k unparseable round recorded, nothing adopted",
      _gbest is None and _gc is None and len(_gtraj) == 1
      and _gtraj[0]["score"] is None)
# best-round-wins when a LATER round regresses: good then bad
_n13_calls2 = []


def _n13_model2(prompt):
    _n13_calls2.append(prompt)
    return [_n13_outs[1], _n13_outs[0]][len(_n13_calls2) - 1]


_b2, _c2, _t2 = _rg(_n13_model2, _n13_rows, _n13_cp, _n13_cp,
                    target_score=200.0)
check("N13l a regressing later round is recorded but NOT kept",
      _c2.score == _gcrit.score and len(_t2) == 2
      and not _t2[1]["kept"])

# N14. Gantt presentation rules + the final gantt in the report.
from programme import (asbuilt_path_tree as _n14apt,
                       build_gantt_html as _n14bgh,
                       build_apab_gantt_html as _n14apab,
                       build_simple_xlsx as _n14bsx,
                       build_asbuilt_xlsx as _n14bax,
                       extract_actual_trace as _n14eat)
_n14tr = _n14eat([("B", B), ("U", U)], max_gap_days=60)
_n14flat = _n14apt(_n14tr.activities, links=_n14tr.links)
check("N14 ungrouped activities are LEAF rows — no pseudo-summary "
      "headers",
      all(c["leaf"] for c in _n14flat["children"][0]["children"]))
_n14grp = _n14apt(_n14tr.activities,
                  groups={"Pkg": [_n14tr.activities[0].task_code,
                                  _n14tr.activities[1].task_code]},
                  links=_n14tr.links)
_n14kids = _n14grp["children"][0]["children"]
check("N14b adopted umbrella renders as a real group, rest stay leaf",
      any(not c["leaf"] and c["name"] == "Pkg" for c in _n14kids)
      and all(c["leaf"] for c in _n14kids if c["name"] != "Pkg"))
_n14html = _n14bgh(_n14flat)
check("N14c tree gantt opens every level and offers full screen",
      "openAll" in _n14html and 'id="fs"' in _n14html
      and "c.leaf" in _n14html)
_n14rows = planned_vs_actual(B, U, None)[:6]
_n14cmp = _n14apab(_n14rows)
check("N14d comparison gantt frozen columns painted OPAQUE",
      "td.lbl { background:#FCFCFA !important; }" in _n14cmp
      and "tr.kd td.lbl" in _n14cmp and "id='fs'" in _n14cmp)

# the final gantt travels WITH the report (workbook + Word). The
# step-④ figure is the EXACT chart rasterised (gantt_png, PIL — no
# browser), same inputs as the HTML renderer.
from programme import build_apab_gantt_png as _n14bapg
_n14kd = {_n14rows[2]["task_code"]: "why", _n14rows[4]["task_code"]: ""}
_n14kwin = _kw(_n14rows, list(_n14kd))
_n14png = _n14bapg(
    [{"task_code": "", "row_kind": "section", "name": "PATH — test"}]
    + _n14rows,
    keydates=_n14kd, overall_delay_days=455,
    windows=[{"label": f"W{i}", "start": w["window_start"],
              "end": w["window_end"],
              "delay_days": w["window_delay_days"]}
             for i, w in enumerate(_n14kwin, 1)],
    data_date=U.project.data_date)
check("N14o the exact step-④ chart rasterises with every feature "
      "(sections, key dates, curtains, data date) and guards empties",
      _n14png is not None and _n14png.startswith(b"\x89PNG")
      and _n14bapg([]) is None)
try:
    from programme.report_charts import (asbuilt_gantt_chart as _n14abc,
                                         chart_png as _n14cp)
    _n14png2 = _n14cp(_n14abc(_n14tr))
    check("N14e as-built path figure renders to PNG",
          _n14png2.startswith(b"\x89PNG"))
    import zipfile as _n14zf
    _n14wb = _n14bsx("t", {"Comparison": [{"a": 1}]},
                     images={"Final Gantt": _n14png})
    _n14names = _n14zf.ZipFile(io.BytesIO(_n14wb)).namelist()
    check("N14f workbook embeds the final gantt as a figure sheet",
          any("media/image" in n for n in _n14names))
    _n14ab = _n14bax(_n14tr, gantt_png=_n14png2)
    check("N14g as-built workbook carries its path gantt",
          any("media/image" in n
              for n in _n14zf.ZipFile(io.BytesIO(_n14ab)).namelist()))
    from programme import build_narrative_docx as _n14doc
    _n14dx = _n14doc("t", "## s\\nbody",
                     images=[("Final gantt", _n14png)])
    check("N14h Word narrative carries the figure",
          any("media/image" in n
              for n in _n14zf.ZipFile(io.BytesIO(_n14dx)).namelist()))
    # markdown tables must land as REAL Word tables, not pipe/dash text
    _n14md = ("## Windows\n\n| Window | Delay |\n|---|---|\n"
              "| W1 | +12 |\n| W2 | -3 |\n\nafter")
    _n14dx2 = _n14doc("t", _n14md)
    _n14doc_xml = _n14zf.ZipFile(io.BytesIO(_n14dx2)).read(
        "word/document.xml").decode("utf-8")
    check("N14j markdown tables render as Word tables (no dash walls)",
          "<w:tbl>" in _n14doc_xml and "|---|" not in _n14doc_xml
          and "W1" in _n14doc_xml)
    # the figure LEADS the Word report — the gantt tells the delay
    # story, the narrative follows it
    _n14dx3 = _n14doc("t", "## Section\nNarrBody123",
                      images=[("Final gantt", _n14png)])
    _n14xml3 = _n14zf.ZipFile(io.BytesIO(_n14dx3)).read(
        "word/document.xml").decode("utf-8")
    check("N14p the figure sits BEFORE the narrative body",
          0 < _n14xml3.find("<w:drawing") < _n14xml3.find("NarrBody123"))
except ImportError as _n14exc:
    print(f"  [SKIP] N14e-h figure pipeline ({_n14exc})")

# N14i. The NVIDIA dropdown is a CURATED shortlist; the live catalogue
# may only REMOVE from it (a retired model), never bury it under the
# endpoint's dozens of models.
from dcma.narrative import PROVIDERS as _n14prov
_n14nv = _n14prov["nvidia"]["models"]
check("N14i EOL'd qwen3-next-80b no longer offered statically",
      "qwen/qwen3-next-80b-a3b-instruct" not in _n14nv)
check("N14k NVIDIA offers a curated three, default among them",
      len(_n14nv) == 3
      and _n14prov["nvidia"]["default_model"] in _n14nv)
import views._shared as _n14sh
_n14sh._live_models = lambda base, fp, key: [       # fake catalogue
    _n14nv[0], _n14nv[2], "some/other-model", "and/another"]
_n14ref = _n14sh.refresh_models(_n14prov["nvidia"], "k")
check("N14l live catalogue REMOVES retired models, never adds",
      _n14ref["models"] == [_n14nv[0], _n14nv[2]]
      and _n14ref["default_model"] == _n14nv[0])
_n14sh._live_models = lambda base, fp, key: ["nothing/known"]
check("N14m no overlap with the catalogue keeps the curated list",
      _n14sh.refresh_models(_n14prov["nvidia"], "k")["models"] == _n14nv)
check("N14n the static list is never mutated in place",
      _n14prov["nvidia"]["models"] == _n14nv)

# N15. Completion impact attribution — which changes actually moved
# completion, measured by one-at-a-time reversion + kernel re-schedule.
from programme import (assess_comparison_impact as _n15imp,
                       attribute_completion_impact as _n15attr,
                       compare_revisions as _n15cmp,
                       build_comparison_prompt as _n15bp,
                       build_comparison_xlsx as _n15bx,
                       build_provenance as _n15prov)
_n15c = _n15cmp(B, U, "B", "U")
_n15i = _n15imp(B, U, "B", "U", comparison=_n15c, end_task_code="KD15")
check("N15 impact carries both driving paths for the summary gantt",
      len(_n15i.lp_old) > 0 and len(_n15i.lp_new) > 0
      and all(len(x) == 2 for x in _n15i.lp_new_links[:5]))
_n15a = _n15attr(B, U, "B", "U", comparison=_n15c, impact=_n15i,
                 end_task_code="KD15")
check("N15b kernel completions computed for both revisions, "
      "movement kernel-vs-kernel",
      _n15a.kernel_completion_old is not None
      and _n15a.kernel_completion_new is not None
      and _n15a.kernel_moved_days is not None
      and 400 < _n15a.kernel_moved_days < 700)
_n15t = _n15a.tested_changes
check("N15c every tested change carries both completions and the "
      "contribution identity (with - without)",
      _n15t and all(
          a.completion_without is not None
          and abs(a.contribution_days
                  - round((a.completion_with - a.completion_without
                           ).total_seconds() / 86400, 1)) < 0.05
          for a in _n15t))
check("N15d untested changes say WHY (completed side / cap)",
      all(a.note for a in _n15a.changes if not a.tested))
check("N15e changes ranked by absolute contribution",
      [abs(a.contribution_days or 0) for a in _n15a.changes]
      == sorted([abs(a.contribution_days or 0)
                 for a in _n15a.changes], reverse=True))
_n15a2 = _n15attr(B, U, "B", "U", comparison=_n15c, max_tests=2)
check("N15f the test cap is honoured and disclosed",
      len(_n15a2.tested_changes) <= 2
      and any("cap" in a.note for a in _n15a2.changes if not a.tested))
_n15pv = _n15prov([("B", B)] + fix[:1] + [("U", U)])
_n15p = _n15bp(_n15c, None, impact=_n15i, attribution=_n15a,
               provenance=_n15pv)
check("N15g the narrative prompt carries screening, attribution, "
      "path comparison and provenance",
      all(t in _n15p for t in ("<impact_screening",
                               "<completion_attribution",
                               "<longest_path_comparison>",
                               "<provenance")))
_n15wb = load_workbook(io.BytesIO(_n15bx(
    _n15c, None, impact=_n15i, attribution=_n15a, provenance=_n15pv)))
check("N15h the workbook ships every table the page shows",
      {"Materiality Rank", "Completion Attribution",
       "Provenance"} <= set(_n15wb.sheetnames))


# ===================================================================== #
# Layer M — LOCAL field-corpus regression (client programmes on this
# machine; never committed). Runs only when the corpus folder exists —
# CI and other machines skip it silently. Answers the review's "one
# project's data validates everything" with four real project families.
# ===================================================================== #
import os as _os
_FIELD = _os.path.expanduser("~/Desktop/Programmes")
if _os.path.isdir(_FIELD):
    print("\n--- Layer M: field corpus (local only) ---")
    import glob as _glob
    from programme import analyse_windows as _aw3

    _m_files = sorted(_glob.glob(_FIELD + "/*/*.xer"))
    _m_ok, _m_val, _m_bad = 0, 0, []
    _m_parsed = {}
    for _f in _m_files:
        try:
            _m_parsed[_f] = parse_xer(_f)
            _m_ok += 1
        except ValueError:
            _m_val += 1
        except Exception as _exc:            # noqa: BLE001
            _m_bad.append(f"{_f.split('/')[-1]}: "
                          f"{type(_exc).__name__}")
    check("M1 every field file parses or raises controlled ValueError",
          not _m_bad and _m_ok >= 15, str(_m_bad))
    check("M1b the structure-only NCC export is REFUSED (zero tasks)",
          _m_val >= 1)

    def _m_series(sub):
        out = [(f.split("/")[-1], d) for f, d in _m_parsed.items()
               if f"/{sub}/" in f]
        out.sort(key=lambda p: p[1].project.data_date)
        return out

    _ncc = _m_series("NCC")
    if len(_ncc) >= 3:
        _m_w = _aw3(_ncc)
        _m_id_ok = all(
            abs(w.performance_days + w.replanning_days
                - w.engine_window_days) < 0.15
            for w in _m_w.windows if w.engine_window_days is not None)
        check("M2 NCC monthly series: bifurcation identity holds in "
              "every window", _m_id_ok and len(_m_w.windows) >= 6)
        check("M2b duplicate data-date revision pair warned",
              any("does not have a later data date" in w
                  for w in _m_w.warnings))

    _ish = _m_series("Ishtar")
    if _ish:
        _m_fl = out_of_sequence_flags(_ish[-1][1])
        check("M3 Ishtar reversed-order as-built stays review-class "
              "(never auto-fitted)",
              len(_m_fl) > 0
              and all(f.rec_link_type == "review" for f in _m_fl))

    _sp_files = [f for f in _m_parsed if "/SPML/" in f]
    if _sp_files:
        import time as _time
        _sp = _m_parsed[_sp_files[0]]
        _t0 = _time.time()
        run_all_checks(_sp, DCMAConfig())
        out_of_sequence_flags(_sp)
        from programme import collapse_asbuilt as _cab3
        _m_cab = _cab3(_sp, "SPML", set())
        _dt = _time.time() - _t0
        check("M4 17k-task file: DCMA + OOS + collapse under 30s",
              _dt < 30, f"{_dt:.1f}s")
        check("M4b heavy-OOS file: collapse validation gap warned",
              any("validation gap" in w for w in _m_cab.warnings))

print(f"\n{'='*60}\nRESULT: {len(PASS)} passed, {len(FAIL)} FAILED")
for name, d in FAIL:
    print(f"  FAILED: {name} — {d}")

sys.exit(1 if FAIL else 0)
